#⚙️ 🟢 FONDATIONS
#01--IMPORTS CORE

import os
import base64
import requests
import urllib.parse
import unicodedata
import re
import time
import html as html_lib
from datetime import datetime
from openpyxl import Workbook

#02--IMPORTS WEB
from flask import Flask, request, redirect, send_file

#03--CONFIG GLOBALE

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FLAG_PATH = os.path.join(BASE_DIR, "no_cleanup.flag")
DB_PATH = os.path.join(BASE_DIR, "films.db")
TMP_PATH = os.path.join(BASE_DIR, "backup_temp.db")
CLEANUP_FILE = os.path.join(BASE_DIR, "last_cleanup.txt")
LAST_BACKUP_FILE = os.path.join(BASE_DIR, "last_backup.txt")

COL = {
    "EMPLACEMENT": 1,
    "TYPE": 3,
    "DISC_ID": 5,   # ✅ colonne E
    "TITRE": 6,
    "ALLOCINE": 7,
    "TMDB_ID": 13,
    "JAQUETTE": 14,
    "ANNEE": 15,
    "GENRES": 16,
    "RESUME": 17,
    "CASTING": 18,
    "ORDRE": 11
}

#04 — ENV / TOKENS

ENV = os.getenv("ENV", "DEV")
TMDB_API_KEY = os.getenv("TMDB_API_KEY")




#🔐 🟡 GITHUB / BACKUP CORE
#05 — GITHUB UTILS

def get_github_token():
    token = os.getenv("GITHUB_TOKEN")
    if not token:
        print("❌ GITHUB_TOKEN manquant")
    return token
    
GITHUB_TOKEN = get_github_token()

#06 — CLEANUP LOGIC

def get_last_backup_delay():

    if not os.path.exists(LAST_BACKUP_FILE):
        return "jamais"

    try:
        with open(LAST_BACKUP_FILE, "r") as f:
            last = float(f.read())
    except:
        return "inconnu"

    delay = int(time.time() - last)

    if delay < 60:
        return f"{delay}s"

    minutes = delay // 60
    if minutes < 60:
        return f"{minutes} min"

    hours = minutes // 60
    if hours < 24:
        return f"{hours} h"

    days = hours // 24
    return f"{days} j"

def should_cleanup():
    if not os.path.exists(CLEANUP_FILE):
        return True

    try:
        with open(CLEANUP_FILE, "r") as f:
            last = float(f.read())
    except Exception:
        return True

    return time.time() - last > 3600


def mark_cleanup():
    with open(CLEANUP_FILE, "w") as f:
        f.write(str(time.time()))
        
#07 — RESTORE DB

def get_latest_backup(files, headers):

    latest_file = None
    latest_date = None

    for f in files:

        try:
            url = f["git_url"]
            r = requests.get(url, headers=headers, timeout=5)

            if r.status_code != 200:
                continue

            data = r.json()

            # 🔐 sécurisation accès date
            date = data.get("committer", {}).get("date")

            if not date:
                continue

            if not latest_date or date > latest_date:
                latest_date = date
                latest_file = f

        except Exception as e:
            print("⚠️ erreur fichier backup :", e)
            continue

    return latest_file


def restore_db():
    
    print("🔥 RESTORE déclenché")

    try:
        token = GITHUB_TOKEN
        repo = "bruno-lille/repo"

        if not token or not repo:
            print("❌ Variables GitHub manquantes")
            return

        headers = {
            "Authorization": f"token {token}"
        }

        # 📥 récupérer les backups
        url = f"https://api.github.com/repos/{repo}/contents/backups"
        r = requests.get(url, headers=headers, timeout=5)

        if r.status_code != 200:
            print("❌ Impossible de récupérer les backups")
            return

        files = r.json()

        # 🔥 garder uniquement les .db
        db_files = [f for f in files if f["name"].endswith(".db")]

        if not db_files:
            print("❌ Aucun backup .db trouvé")
            return

        latest = sorted(db_files, key=lambda x: x["name"], reverse=True)[0]

        download_url = latest["download_url"]

        r = requests.get(download_url, timeout=5)

        if r.status_code != 200:
            print("❌ Erreur téléchargement DB")
            return

        # 🔐 sécurité : ne pas écraser une DB valide
        if os.path.exists(DB_PATH):
            size = os.path.getsize(DB_PATH)

            if size > 10000:  # seuil simple (DB valide)
                print("⚠️ DB locale déjà valide → restore ignoré")
                return

        # 💾 écriture directe (PAS de vérification locale)
        tmp_restore = DB_PATH + ".restore"

        with open(tmp_restore, "wb") as f:
            f.write(r.content)

        # 🔍 vérification taille minimale
        if os.path.getsize(tmp_restore) < 1000:
            print("❌ Restore invalide (fichier trop petit)")
            os.remove(tmp_restore)
            return

        # 🔥 remplacement sécurisé
        os.replace(tmp_restore, DB_PATH)

        print(f"✅ DB restaurée : {latest['name']}")

    except Exception as e:
        print("❌ ERREUR RESTORE :", e)

def init_app():
    if ENV == "PROD":
        print("🌐 Mode PROD → vérification DB")

        if not os.path.exists(DB_PATH):
            print("📥 DB absente → restauration GitHub")
            restore_db()
        else:
            print("✅ DB déjà présente → OK")

    else:
        print("🧪 Mode DEV → pas de restore")
        
#💾 🔵 DATABASE
#08 — SQL CORE
#Ancien03------------FONCTION SQL------------

def search_films_sql(query, exact_mode=False):
    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    q = query.lower()

    if exact_mode:
        cursor.execute(
            "SELECT * FROM films WHERE LOWER(titre) = ?",
            (q,)
        )
    else:
        cursor.execute("SELECT * FROM films")

    results = cursor.fetchall()
    conn.close()

    return results
    
#09 — LOAD DATA

def load_data():
    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM films
    ORDER BY
        CASE WHEN ordre IS NULL THEN 999999 ELSE ordre END ASC,
        annee ASC
    """)
    rows = cursor.fetchall()

    titres, liens = [], []
    col_A, col_C = [], []
    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []
    disc_ids, ordres = [], []

    for row in rows:
        disc_ids.append(row[1])
        col_A.append(row[2])
        col_C.append(row[3])
        titres.append(row[4])
        liens.append(row[5])
        tmdb_ids.append(row[6])
        jaquettes.append(row[7])
        annees.append(row[8])
        genres.append(row[9])
        resumes.append(row[10])
        casting.append(row[11])
        ordres.append(row[12])

    conn.close()

    return titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids, ordres
    
DATA = None
    
def load_data_cached():
    global DATA

    if DATA is None:
        DATA = load_data()

    return DATA
    
#🌐 🟣 API EXTERNES
#10 — TMDB SEARCH

def search_tmdb_multi(query):
    print("TMDB KEY:", TMDB_API_KEY)
    print("QUERY:", query)
    print("QUERY TMDB:", query)
    print("API KEY:", TMDB_API_KEY)
    query = query.strip()

    results = []

    def fetch(lang):
        try:
            url = "https://api.themoviedb.org/3/search/movie"
            params = {
                "api_key": TMDB_API_KEY,
                "query": query,
                "language": lang,
                "region": "FR",
                "include_adult": False
            }

            data = requests.get(url, params=params, timeout=5).json()

            for film in data.get("results", []):
                poster = film.get("poster_path")
                img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""

                results.append({
                    "id": film["id"],
                    "title": film["title"],
                    "year": film.get("release_date", "")[:4],
                    "img": img
                })
        except:
            pass

    # 🔥 recherche FR
    fetch("fr-FR")

    # 🔥 fallback anglais
    if len(results) < 5:
        fetch("en-US")

    # 🔥 suppression doublons
    unique = {film["id"]: film for film in results}

    # 🔥 tri par année décroissante
    sorted_results = sorted(
        unique.values(),
        key=lambda x: (x["year"] or "0", x["title"]),
        reverse=True
    )

    return sorted_results[:5]
    
#11 — TMDB HELPERS

def extract_tmdb_id(value):

    value = value.strip()

    # cas simple : nombre direct
    if value.isdigit():
        return value

    # cas URL TMDB (robuste)
    match = re.search(r'/movie/(\d+)', value)
    if match:
        return match.group(1)

    return None
    
#🧰 🟠 UTILITAIRES
#12 — NORMALISATION

def normalize(text):
    if not text:
        return ""

    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")

    text = re.sub(r'[^a-z0-9]', '', text)

    return text
    
def normalize_words(text):
    if not text:
        return []

    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")

    # garder les mots séparés
    words = re.findall(r'\b[a-z0-9]+\b', text)

    return words
    
#13 — UTILS GENERIQUES

def short_text(text, max_len=140):
    return text if len(text) <= max_len else text[:max_len] + "..."
    
def safe_int(val, default=0):
    try:
        return int(val)
    except:
        return default
        

        
#14 — EXCEL UTILS
#ancien06 ----------- EXTRACTION ID TMDB -----------


def generate_disc_id(ws):

    max_id = 0

    for row in ws.iter_rows(min_row=2):

        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val.startswith("DISC-"):
            try:
                num = int(val.replace("DISC-", ""))
                if num > max_id:
                    max_id = num
            except:
                pass

    new_id = max_id + 1

    return f"DISC-{str(new_id).zfill(5)}"
    
def find_row_by_disc_id(ws, disc_id):

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val == disc_id:
            return idx

    return None
    


        
#🎨 ⚪ UI / FRONT
#15 — STYLE

def get_style():
    return """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    
    <script>
    window.onload = function() {
        let input = document.getElementById("title");
        if (input) {
            input.focus();
            input.select();
        }
    }
    </script>
  
    <script>
    async function smartPaste(input) {

        // sélectionner tout le texte
        input.select();

        try {
            // tentative de lecture du presse-papier
            const text = await navigator.clipboard.readText();

            if (text && text.startsWith("http")) {
                input.value = text;
            }

        } catch (err) {
            // fallback → rien (iPhone va juste sélectionner)
        }
    }
    </script>
    
    <script>
    function updateCount() {

        let q = document.getElementById("search").value;

        if (q.length === 0) {
            document.getElementById("count").innerText = "";
            return;
        }

        fetch("/count?q=" + encodeURIComponent(q))
            .then(res => res.text())
            .then(data => {
                document.getElementById("count").innerText =
                    data + " résultats";
            });
    }
    </script>
    
    <script>
    function setTMDB(id) {
        const input = document.getElementById("tmdb_input");

        if (!input) {
            alert("Champ TMDB introuvable");
            return;
        }

        input.value = id;
        input.focus();

        window.scrollTo({ top: 0, behavior: "smooth" });
    }

    function setTitle(title) {
        const input = document.getElementById("title");

        if (!input) {
            alert("Champ titre introuvable");
            return;
        }

        input.value = title;
        input.focus();
        input.style.background = "#2a9df4";

        setTimeout(() => {
            input.style.background = "";
        }, 300);

        window.scrollTo({ top: 0, behavior: "smooth" });
    }
    </script>
    
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}

    input {
        padding:14px;
        width:100%;
        font-size:16px;
        border-radius:10px;
        border:none;
        text-align:center;
        margin-bottom:10px;
    }

    .card {
        background:#1c1c1c;
        padding:15px;
        border-radius:16px;
        margin:15px auto;
        width:95%;
        max-width:420px;
    }

    img {width:100%; border-radius:12px;}

    .titre a {
        color:#4fc3f7;
        font-size:18px;
        font-weight:bold;
        text-decoration:none;
    }

    .btn-row {
        display:flex;
        gap:8px;
        margin-top:12px;
    }

    .btn {
        flex:1;
        padding:12px;
        border-radius:10px;
        text-decoration:none;
        font-weight:bold;
        font-size:14px;
    }
    
    .btn:hover {
        opacity:0.85;
    }

    .allocine {
        background:#2a9df4;
        color:white;
        font-weight:bold;
    }
    .new {background:#e50914; color:white;}
    .update {
        background:#ffb74d;
        color:black;
    }
    .retour {background:#444; color:white;}
    </style>
    """

#16 - composants UI réutilisables
nav_buttons = """
<div class="card">
    <div class="btn-row">
        <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
        <a class="btn new" href="/">🏠 Accueil</a>
    </div>
</div>
"""


#🚀 🔴 APP FLASK
#17 — INIT APP

app = Flask(__name__)

APP_VERSION = "V1-dev"
APP_BUILD = "2026-05-04_14-30-16"
APP_NOTE = "dev en cours"



#18 — CACHE reset



def reset_cache():
    global DATA
    DATA = None
  
#🔍 🟢 ROUTES PRINCIPALES  
#19 — HOME
@app.route("/", methods=["GET"])
def home():

    show_suggest = request.args.get("suggest")
    print("SUGGEST =", show_suggest)

    titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids, ordres = load_data_cached()

    query_raw = request.args.get("q", "").strip()
    mode = request.args.get("mode")

    exact_mode = False

    if query_raw.startswith("(") and query_raw.endswith(")"):
        exact_mode = True
        query = query_raw[1:-1].strip()
    else:
        query = query_raw

    html = get_style()

    # =========================
    # 🔥 PAGE ACCUEIL
    # =========================
    if not query:

        html += f"""
        <div style="
            position:fixed;
            top:8px;
            right:10px;
            font-size:12px;
            color:white;
            background:#222;
            padding:4px 8px;
            border-radius:6px;
            z-index:9999;
        ">
            {APP_VERSION} | {APP_BUILD}
        </div>
        
        <div style="
            position:fixed;
            bottom:20px;
            right:10px;
            font-size:12px;
            color:#aaa;
            background:#222;
            padding:6px 10px;
            border-radius:8px;
            z-index:9999;
        ">
            🕒 {get_last_backup_delay()}
        </div>

        <h1>🎬 Ma vidéothèque</h1>
      

        <form>

            <div id="count" style="font-size:20px; margin-bottom:10px;"></div>

            <input name="q" id="search" placeholder="Tape un film..." onkeyup="updateCount()">

            <div class="card">
                <div class="btn-row">
                    <button class="btn allocine" name="mode" value="fast">🔎 Recherche</button>
                    <button class="btn new" name="mode" value="poster">🎬 Jaquettes</button>
                </div>
            </div>
        </form>
   
        """
        html += f"""
        <div class="card">

            <div style="display:flex; gap:8px;">

                <a class="btn allocine" href="/download_all" style="flex:1;">
                    💾 Télécharger
                </a>

                <form action="/upload_db" method="post" enctype="multipart/form-data" style="flex:1; margin:0;">

                    <input type="file" name="file" accept=".db" style="width:100%; margin-bottom:5px;">

                    <button class="btn update" style="width:100%;"
                        onclick="return confirm('⚠️ Remplacer la base actuelle ? Cette action est irréversible.')"
                        📥 Restaurer
                    </button>

                </form>

            </div>

        </div>
        """
        return html

    # =========================
    # 🔍 RECHERCHE
    # =========================
    rows = search_films_sql(query, exact_mode)
    results = list(rows)

    # 🔥 fallback large si vide
    if not results:
        rows = search_films_sql("", False)
        results = list(rows)

    # 🔥 normalisation
    # 🔥 NORMALISATION
    q_norm = normalize(query)

    # 🔥 découpage AVANT normalisation
    q_words_raw = query.lower().split()

    # 🔥 mots normalisés individuellement
    q_words = [normalize(w) for w in q_words_raw]

    filtered = []

    for row in results:

        titre_norm = normalize(row["titre"])

        # 🔥 MODE EXACT
        if exact_mode:
            if titre_norm == q_norm:
                filtered.append(row)
            continue

        # 🔥 MODE MULTI-MOTS INTELLIGENT
        match = True

        for q in q_words:
            if q not in titre_norm:
                match = False
                break

        if match:
            filtered.append(row)

    results = [r for r in filtered if r["titre"]]

    # =========================
    # 🎯 SI RESULTATS
    # =========================
    if results:

        bloc = f"<h2>🎬 {query} {len(results)} résultat(s)</h2>"
        bloc += nav_buttons

        for row in results:

            poster_html = f'<img src="{row["jaquette"]}">' if row["jaquette"] else ""

            bloc += f"""
            <div class="card">
                {poster_html}

                <div class="titre">
                    <div style="color:#888; font-size:12px; margin-bottom:4px;">
                        🆔 {row["disc_id"]}
                    </div>

                    <a href="https://www.themoviedb.org/movie/{row["tmdb_id"]}" target="_blank">
                        {row["titre"]}
                    </a>
                </div>

                <div style="color:#b0b0b0; font-size:16px; margin-top:6px;">
                    📀 {row["type"]} &nbsp;&nbsp;
                    📁 {row["emplacement"]} &nbsp;&nbsp;
                    📅 {row["annee"]}
                </div>

                <div>{row["genres"]}</div>
                <div>{row["casting"]}</div>

                <div class="btn-row">
                    <a class="btn allocine" href="{row["allocine"]}" target="_blank">Allociné</a>
                    <a class="btn update" href="/suggest_update/{row["disc_id"]}?q={urllib.parse.quote(query_raw)}">🔄 Corriger</a>
                    <a class="btn new" href="/delete/{row["disc_id"]}" onclick="return confirm('Supprimer ce film ?')">
                        🗑 Supprimer
                    </a>
                </div>
            </div>
            """
        query_encoded = urllib.parse.quote(query_raw)

        bloc += f"""
        <div class="card">
            <a class="btn allocine" href="/?q={query_encoded}&mode={mode}&suggest=1">
                🔎 Plus de résultats TMDB
            </a>
        </div>
        """
        
        # 🔥 SUGGESTIONS TMDB
        if show_suggest:
            results_tmdb = search_tmdb_multi(query_raw)

            bloc += "<h2>🎬 Suggestions TMDB</h2>"

            for film in results_tmdb:
                bloc += f"""
                <div class="card">
                    <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
                    <div>{film['title']} ({film['year']})</div>

                    <div class="btn-row">
                        <a class="btn new" href="/add/{film['id']}?q={urllib.parse.quote(query_raw)}">
                            ➕ Ajouter
                        </a>
                    </div>
                </div>
                """

        bloc += nav_buttons
        return html + bloc
        
    # 🔥 AUCUN RESULTAT → TMDB DIRECT
    results_tmdb = search_tmdb_multi(query_raw)

    # =========================
    # 🎬 SI TMDB TROUVE
    # =========================
    if results_tmdb:

        bloc = f"<h2>🎬 {query} n'existe pas dans ma liste mais voici les résultats TMDB</h2>"
        bloc += nav_buttons

        for film in results_tmdb:
            bloc += f"""
            <div class="card">
                <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
                <div>{film['title']} ({film['year']})</div>

                <div class="btn-row">
                    <a class="btn new" href="/add/{film['id']}?q={urllib.parse.quote(query_raw)}">
                        ➕ Ajouter
                    </a>
                </div>
            </div>
            """

        bloc += nav_buttons
        return html + bloc

    # =========================
    # ❌ SI TMDB VIDE → FORMULAIRE
    # =========================
    query_encoded = urllib.parse.quote(query_raw)

    bloc = f"<h2>❌ {query} introuvable — Ajout manuel</h2>"
    bloc += nav_buttons

    bloc += f"""
    <div class="card">
        <h3>✍️ Ajouter ce film</h3>

        <form method="post" action="/manual_add">

            <input name="title" value="{query.title()}" placeholder="Titre" required><br><br>

            <input name="emplacement" placeholder="📁 Emplacement"><br><br>

            <input name="type" placeholder="📀 Type (DVD/BLURAY)"><br><br>

            <input name="allocine" value="https://www.google.com/search?q={query_encoded}+allocine" placeholder="🔗 Allociné"><br><br>

            <input name="tmdb_input" placeholder="ID ou lien TMDB"><br><br>

            <div class="btn-row">
                <a class="btn allocine" href="https://www.google.com/search?q={query_encoded}+allocine+film" target="_blank">
                    🎬 Chercher Allociné
                </a>

                <a class="btn allocine" href="https://www.themoviedb.org/search?query={query_encoded}" target="_blank">
                    🔎 Chercher TMDB
                </a>
            </div>

            <br>

            <button class="btn new">➕ Ajouter</button>

        </form>
    </div>
    """

    bloc += nav_buttons
    return html + bloc
    
#20 — COUNT
@app.route("/count")
def count():

    query_raw = request.args.get("q", "").strip()

    if not query_raw:
        return "0"

    # 🔥 gestion mode exact (parenthèses)
    exact_mode = False
    if query_raw.startswith("(") and query_raw.endswith(")"):
        exact_mode = True
        query = query_raw[1:-1].strip()
    else:
        query = query_raw

    # 🔥 récupération des données
    rows = search_films_sql(query, exact_mode)
    results = list(rows)

    # 🔥 fallback large
    if not results:
        rows = search_films_sql("", False)
        results = list(rows)

    # 🔥 NORMALISATION IDENTIQUE À home()
    q_norm = normalize(query)
    q_words = [normalize(w) for w in query.lower().split()]

    filtered = []

    for row in results:

        titre_norm = normalize(row["titre"])

        # 🔥 mode exact
        if exact_mode:
            if titre_norm == q_norm:
                filtered.append(row)
            continue

        # 🔥 multi-mots intelligent
        match = True
        for q in q_words:
            if q not in titre_norm:
                match = False
                break

        if match:
            filtered.append(row)

    return str(len(filtered))
    
#✏️ 🟡 GESTION FILMS
#21 — SUGGEST UPDATE
@app.route("/suggest_update/<disc_id>")
def suggest_update(disc_id):

    import sqlite3
    import urllib.parse
    
    print("DISC_ID REÇU =", disc_id)

    query = request.args.get("q", "")
    query_encoded = urllib.parse.quote(query)
    

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()

    
    if film:
        print("FILM TROUVÉ =", film["titre"])
    else:
        print("❌ AUCUN FILM TROUVÉ")

    conn.close()

    if not film:
        return "❌ Film introuvable"
        
    results_tmdb = search_tmdb_multi(query)

    title = film["titre"] or ""
    emplacement = film["emplacement"] or ""
    type_disc = film["type"] or ""
    allocine = film["allocine"] or ""
    tmdb_id = film["tmdb_id"] or ""
    ordre = film["ordre"] or ""

    html = get_style()

    html += f"""
    <div style="
        position:fixed;
        top:10px;
        right:15px;
        font-size:12px;
        color:#888;
        z-index:999;
    ">
        {APP_VERSION} | {APP_BUILD}
    </div>
    """

    html += f"""
    <h2>🛠 Modification / Mise à jour</h2>

    <div class="card">
        <a class="btn retour" href="/?q={query}">⬅️ Retour</a>
        <a class="btn new" href="/">❌ Annuler</a>
    </div>

    <div class="card">
        <h3>✏️ Modifier complètement</h3>

        <form action="/manual_update/{disc_id}?q={query_encoded}" method="post">
        
            
            <div style="margin-bottom:10px; font-size:13px; color:#888;">
                🆔 {film["disc_id"]}
            </div>

            <div style="font-size:11px; color:#555;">
                DEBUG: {film["disc_id"]}
            </div>

            <input type="hidden" name="disc_id" value="{film['disc_id']}">

            <input id="title" name="title" value="{title}" placeholder="Titre" autofocus style="margin-bottom:10px;">

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement" style="margin-bottom:10px;">

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)" style="margin-bottom:10px;">

            <div style="display:flex; align-items:center; gap:6px; margin-bottom:10px;">

                <input name="allocine"
                       value="{allocine}"
                       placeholder="🔗 Allociné"
                       style="flex:1;"
                       onclick="smartPaste(this)">

                <a href="https://www.google.com/search?q={urllib.parse.quote(title)}+allocine"
                   target="_blank"
                   style="
                       font-size:12px;
                       padding:6px 10px;
                       background:#2a9df4;
                       color:white;
                       border-radius:6px;
                       text-decoration:none;
                       height:42px;
                       display:flex;
                       align-items:center;
                       justify-content:center;
                   ">
                   🔎
                </a>

            </div>

            <input name="ordre" value="{ordre}" placeholder="🔢 Ordre" type="number" min="1" style="margin-bottom:10px;">

            <div style="display:flex; align-items:center; gap:6px; margin-bottom:10px;">

                <input id="tmdb_input"
                       name="tmdb_input"
                       value="{tmdb_id}"
                       placeholder="ID ou lien TMDB"
                       style="flex:1;"
                       onclick="smartPaste(this)">

                <a href="https://www.themoviedb.org/search/movie?query={urllib.parse.quote(title)}"
                   target="_blank"
                   style="
                       font-size:12px;
                       padding:6px 8px;
                       background:#e50914;
                       color:white;
                       border-radius:6px;
                       text-decoration:none;
                       height:42px;
                       display:flex;
                       align-items:center;
                       justify-content:center;
                   ">
                   🔎
                </a>

            </div>

            <div style="text-align:center; margin-top:15px;">

                <button type="submit" class="btn update" style="width:70%;">
                    💾 Mettre à jour
                </button>

            </div>

        </form>
    </div>
    """

    for film in results_tmdb:

        title_safe = film["title"].replace("'", "\\'")

        html += f"""
        <div class="card">
            <a href="https://www.themoviedb.org/movie/{film['id']}" target="_blank">
                <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
            </a>

            <div class="titre">{film['title']} ({film['year']})</div>

            <div class="btn-row">

                <button type="button" class="btn update"
                    onclick="window.location.href='/confirm_update/{disc_id}/{film['id']}?q={query_encoded}&title=1'">
                    ⚡ Tout remplir (titre inclus)
                </button>

                <button type="button" class="btn allocine"
                    onclick="window.location.href='/confirm_update/{disc_id}/{film['id']}?q={query_encoded}&title=0'">
                    🧠 Garder le titre actuel
                </button>

            </div>
        </div>
        """
    # print("DISC_ID REÇU =", disc_id)
    # print("FILM TROUVÉ =", film["titre"])
    
    return html
    
#22 — UPDATE AUTO
@app.route("/confirm_update/<disc_id>/<int:tmdb_id>")
def confirm_update(disc_id, tmdb_id, query=""):
    use_title = request.args.get("title") == "1"

    import sqlite3

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {"api_key": TMDB_API_KEY, "language": "fr-FR", "append_to_response": "credits"}

    data = requests.get(url, params=params, timeout=5).json()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    if use_title:
        titre = data.get("title", "")
    else:
        cursor.execute("SELECT titre FROM films WHERE disc_id = ?", (disc_id,))
        row = cursor.fetchone()
        titre = row[0] if row else ""


    cursor.execute("""
    UPDATE films SET
        titre = ?,
        tmdb_id = ?,
        jaquette = ?,
        annee = ?,
        genres = ?,
        resume = ?,
        casting = ?
    WHERE disc_id = ?
    """, (
        titre,
        tmdb_id,
        f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}",
        data.get("release_date", "")[:4],
        ", ".join([g["name"] for g in data.get("genres", [])]),
        data.get("overview", ""),
        ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]]),
        disc_id
    ))

    conn.commit()
    conn.close()

    reset_cache()

    query = request.args.get("q", "")
    return redirect(f"/?q={query}")
    
#23 — UPDATE MANUEL
@app.route("/manual_update/<disc_id>", methods=["POST"])
def manual_update(disc_id):

    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 🔥 vérifie si le film existe
    cursor.execute("SELECT id FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()

    if not film:
        conn.close()
        return "❌ Introuvable"

    # 🔥 mise à jour des champs
    cursor.execute("""
        UPDATE films SET
            titre = ?,
            emplacement = ?,
            type = ?,
            allocine = ?,
            ordre = ?,
            tmdb_id = ?,
            jaquette = ?,
            annee = ?,
            genres = ?,
            resume = ?,
            casting = ?
        WHERE disc_id = ?
    """, (
        request.form.get("title", ""),
        request.form.get("emplacement", ""),
        request.form.get("type", ""),
        request.form.get("allocine", ""),
        int(request.form.get("ordre")) if request.form.get("ordre") else None,
        request.form.get("tmdb_input", ""),
        request.form.get("poster", ""),
        request.form.get("year", ""),
        request.form.get("genres", ""),
        request.form.get("overview", ""),
        request.form.get("cast", ""),
        disc_id
    ))

    conn.commit()
    conn.close()
    reset_cache()

    # 🔥 TMDB optionnel
    raw = request.form.get("tmdb_input", "")
    tmdb_id = extract_tmdb_id(raw) if raw else None

    query = request.args.get("q", "")

    if tmdb_id:
        
        reset_cache()
        return confirm_update(disc_id, int(tmdb_id), query)

    
    return redirect(f"/?q={query}")
    
#24 — ADD FILM
#ancien13 ----------- PAGE AJOUT FILM -----------
@app.route("/add/<int:tmdb_id>")
def add_movie(tmdb_id):

    emplacement = request.args.get("emplacement", "")
    type_disc = request.args.get("type", "")
    allocine = request.args.get("allocine", "")

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": TMDB_API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params, timeout=5).json()

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]
    genres = ", ".join([g["name"] for g in data.get("genres", [])])
    overview = data.get("overview", "")
    poster = f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}"
    cast = ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]])
    query_allocine = urllib.parse.quote(f"{title} {year} allocine fichefilm")

    return get_style() + f"""
    <h2>➕ Ajouter ce film</h2>

    <div class="card">
        <img src="{poster}">
        <h3>{title} ({year})</h3>
        <p>{genres}</p>
        <p>{cast}</p>
        <p>{overview}</p>
    </div>

    <div class="card">
        <form method="post" action="/confirm_add">

            <input type="hidden" name="tmdb_id" value="{tmdb_id}">
            <input type="hidden" name="title" value="{title}">
            <input type="hidden" name="year" value="{year}">
            <input type="hidden" name="genres" value="{genres}">
            <input type="hidden" name="overview" value="{overview}">
            <input type="hidden" name="cast" value="{cast}">
            <input type="hidden" name="poster" value="{poster}">
            
            <input name="title" value="{title}" required><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement" required><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)" required><br><br>



            <input name="allocine" 
                  value="https://www.google.com/search?q={query_allocine}">

            <div class="btn-row">
                <a class="btn allocine" 
                   href="https://www.google.com/search?q={query_allocine}" 
                   target="_blank">
                    🎬 Chercher Allociné
                </a>
            </div>
            <br>

            <button class="btn new">✅ Ajouter</button>
        </form>
    </div>

    <div class="card">
        <a class="btn retour" href="/">⬅️ Retour</a>
    </div>
    """
            
#25 — CONFIRM ADD
@app.route("/confirm_add", methods=["POST"])
def confirm_add():

    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 🔍 DONNÉES
    titre = request.form.get("title", "").strip()
    titre_norm = normalize(titre)

    # 🔍 CHECK DOUBLON
    cursor.execute("SELECT titre, disc_id, type, emplacement FROM films")
    rows = cursor.fetchall()

    doublons = [
        r for r in rows
        if normalize(r[0]) == titre_norm
    ]

    # 🚨 SI DOUBLON → PAGE CONFIRMATION
    if doublons:

        page = get_style()
        page += f"<h2>⚠️ Doublon détecté : {html_lib.escape(titre)}</h2>"
        page += "<p>Ce film existe déjà dans votre collection :</p>"

        for d in doublons:
            page += f"""
        <div style="margin:8px 0; font-size:15px;">
            🎬 {html_lib.escape(d[0])} <br>
            📀 {html_lib.escape(d[2] or '')} &nbsp;&nbsp;
            🆔 {html_lib.escape(d[1] or '')} &nbsp;&nbsp;
            📁 {html_lib.escape(d[3] or '-')}
        </div>
        """

        # 🔥 FORMULAIRE POUR FORCER
        page += f"""
        <div class="card">
            <form method="post" action="/confirm_add_force">

                <input type="hidden" name="title" value="{html_lib.escape(request.form.get('title',''))}">
                <input type="hidden" name="emplacement" value="{html_lib.escape(request.form.get('emplacement',''))}">
                <input type="hidden" name="type" value="{html_lib.escape(request.form.get('type',''))}">
                <input type="hidden" name="allocine" value="{html_lib.escape(request.form.get('allocine',''))}">
                <input type="hidden" name="tmdb_id" value="{html_lib.escape(request.form.get('tmdb_id',''))}">
                <input type="hidden" name="poster" value="{html_lib.escape(request.form.get('poster',''))}">
                <input type="hidden" name="year" value="{html_lib.escape(request.form.get('year',''))}">
                <input type="hidden" name="genres" value="{html_lib.escape(request.form.get('genres',''))}">
                <input type="hidden" name="overview" value="{html_lib.escape(request.form.get('overview',''))}">
                <input type="hidden" name="cast" value="{html_lib.escape(request.form.get('cast',''))}">

                <button class="btn new">✅ Ajouter quand même</button>
            </form>

            <a class="btn retour" href="/">❌ Annuler</a>
        </div>
        """

        conn.close()
        return page

    # ✅ SINON → INSERT NORMAL
    return insert_film(request.form)
    
def insert_film(form):

    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 🔢 DISC_ID AUTO
    cursor.execute("SELECT disc_id FROM films ORDER BY id DESC LIMIT 1")
    last = cursor.fetchone()

    if last and last[0].startswith("DISC-"):
        num = int(last[0].replace("DISC-", ""))
        new_id = f"DISC-{str(num + 1).zfill(5)}"
    else:
        new_id = "DISC-00001"

    cursor.execute("""
    INSERT INTO films (
        disc_id, emplacement, type, titre, allocine,
        tmdb_id, jaquette, annee, genres, resume, casting, ordre
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        new_id,
        form.get("emplacement", ""),
        form.get("type", ""),
        form.get("title", ""),
        form.get("allocine", ""),
        form.get("tmdb_id", ""),
        form.get("poster", ""),
        form.get("year", ""),
        form.get("genres", ""),
        form.get("overview", ""),
        form.get("cast", ""),
        None
    ))

    conn.commit()
    conn.close()

    reset_cache()

    return get_style() + """
    <h2>✅ Film ajouté !</h2>

    <div class="card">
        <a class="btn allocine" href="/">🏠 Retour accueil</a>
    </div>
    """
    
@app.route("/confirm_add_force", methods=["POST"])
def confirm_add_force():
    return insert_film(request.form)
    
#26 — DELETE
@app.route("/delete/<disc_id>")
def delete_movie(disc_id):

    import sqlite3

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("DELETE FROM films WHERE disc_id = ?", (disc_id,))

    conn.commit()
    conn.close()

    reset_cache()

    return """
    <div style="font-size:28px; text-align:center; margin-top:40px;">
        🗑 Film supprimé
    </div>

    <div style="text-align:center; margin-top:20px;">
        <a href="/" style="font-size:20px;">⬅️ Retour</a>
    </div>
    """
    
#27 — PREFILL
@app.route("/prefill/<disc_id>/<int:tmdb_id>")
def prefill(disc_id, tmdb_id):

    import sqlite3

    # 🔥 récupération TMDB
    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": TMDB_API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params, timeout=5).json()

    poster_path = data.get("poster_path")
    poster = f"https://image.tmdb.org/t/p/w300{poster_path}" if poster_path else ""

    genres = ", ".join([g["name"] for g in data.get("genres", [])]) if data.get("genres") else ""
    overview = data.get("overview", "")
    cast = ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]])

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]

    # 🔥 récupération données existantes (SQLite)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()

    conn.close()

    emplacement = film["emplacement"] if film else ""
    type_disc = film["type"] if film else ""
    allocine = film["allocine"] if film else ""
    ordre = film["ordre"] if film else ""

    return get_style() + f"""
    <h2>✨ Pré-remplissage</h2>

    <div class="card">
        <h3>{title} ({year})</h3>
    </div>

    <div class="card">
        <form action="/manual_update/{disc_id}" method="post">
        
            

            <input name="title" value="{title}" placeholder="Titre"><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement"><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)"><br><br>

            <input name="allocine" value="{allocine}" placeholder="🔗 Allociné"><br><br>

            <input name="ordre" value="{ordre}" placeholder="🔢 Ordre" type="number" min="1"><br><br>

            <input id="tmdb_input" name="tmdb_input" value="{tmdb_id}" placeholder="ID ou lien TMDB">

            <button class="btn new">✅ Valider</button>

            <div class="card">
                <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
            </div>

        </form>
    </div>
    """
    
#28 — MANUAL ADD
#ancien20---------Ajout Nanuel fiche vide------------------
@app.route("/manual_add", methods=["POST"])
def manual_add():

    title = request.form.get("title")
    emplacement = request.form.get("emplacement")
    type_disc = request.form.get("type")
    allocine = request.form.get("allocine")
    tmdb_input = request.form.get("tmdb_input", "").strip()

    # 🔥 nettoyage TMDB ID
    import re
    tmdb_id = ""

    if tmdb_input.isdigit():
        tmdb_id = tmdb_input
    elif "themoviedb.org" in tmdb_input:
        match = re.search(r"/movie/(\\d+)", tmdb_input)
        if match:
            tmdb_id = match.group(1)

    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 🔥 génération disc_id automatique
    cursor.execute("SELECT MAX(id) FROM films")
    last_id = cursor.fetchone()[0] or 0
    next_id = last_id + 1
    disc_id = f"DISC-{next_id:05d}"

    # 🔥 INSERT COMPLET
    cursor.execute("""
        INSERT INTO films (disc_id, titre, emplacement, type, allocine, tmdb_id)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (disc_id, title, emplacement, type_disc, allocine, tmdb_id))

    conn.commit()
    conn.close()

    # 🔥 retour à la recherche
    return redirect(f"/?q={title}")
    
#🔁 🔵 SYSTÈME / INFRA
#29 — BACKUP
@app.route("/backup_db", methods=["GET", "HEAD"])
def backup_db():
    if not GITHUB_TOKEN:
        return "❌ GITHUB_TOKEN manquant"
    
    print("🧪 NOUVEAU CLEANUP ACTIF")

    import os
    import base64
    import requests
    import sqlite3
    import shutil


    print("STEP 1: start backup")



    token = GITHUB_TOKEN
    repo = "bruno-lille/repo"

    headers = {
        "Authorization": f"token {token}"
    }

    if not os.path.exists(DB_PATH):
        return "❌ films.db introuvable"

    print("STEP 2: DB found")

    try:
        # 🔥 1. Sécuriser SQLite
        conn = sqlite3.connect(DB_PATH)
        conn.commit()
        conn.execute("PRAGMA wal_checkpoint(FULL);")
        conn.close()

        # 🔥 2. Copier version stable
        shutil.copyfile(DB_PATH, TMP_PATH)
        print("STEP 3: DB copied")

        # 🔥 3. Lire copie
        with open(TMP_PATH, "rb") as f:
            raw = f.read()

        print("STEP 4: DB read OK, size =", len(raw))

        if len(raw) < 1000:
            return "❌ DB invalide"

        content = base64.b64encode(raw).decode()

        # 🔥 4. Upload GitHub
        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"backups/films_{now}.db"

        url = f"https://api.github.com/repos/{repo}/contents/{filename}"

        data = {
            "message": f"backup {now}",
            "content": content,
            "branch": "main"
        }

        r = requests.put(url, json=data, headers=headers)
        backup_status = r.status_code

        print("STEP 5:", r.status_code, r.text)

        if r.status_code not in [200, 201]:
            return f"❌ Backup erreur {r.status_code}"

        print(f"✅ Backup créé : {filename}")
        # 🔥 enregistrer timestamp backup
        with open(LAST_BACKUP_FILE, "w") as f:
            f.write(str(time.time()))
        
        # 🔥 sécurité import
        if os.path.exists(FLAG_PATH):
            print("⚠️ Cleanup désactivé (import récent)")
            os.remove(FLAG_PATH)
            return f"Backup OK → {backup_status}"
        


        # ==================================================
        # 🧠 NETTOYAGE INTELLIGENT
        # ==================================================
        
        

        print("STEP 6: listing backups")

        url = f"https://api.github.com/repos/{repo}/contents/backups"
        r = requests.get(url, headers=headers, timeout=5)

        if r.status_code != 200:
            print("⚠️ Impossible de lister les backups")
            return f"Backup OK → {backup_status}"

        files = r.json()
        print("STEP 7: files received =", len(files))

        db_files = [
            f for f in files
            if f["name"].startswith("films_") and f["name"].endswith(".db")
        ]

        print("STEP 8: db_files =", len(db_files))
        
        # 🧠 ANTI-SPAM CLEANUP
        
        # ✅ RÈGLE D’OR
        #   👉 NE JAMAIS désactiver ce bloc (if not should_cleanup():) sans désactiver aussi le delete
        if not should_cleanup():
            print("⏱️ Cleanup ignoré (moins de 1h)")
            return f"Backup OK → {backup_status}"
            
        print("🧠 Cleanup autorisé → lancement (timestamp enregistré)")
        mark_cleanup()
            

        # 🔥 TRI

        def extract_date(f):
            try:
                name = f["name"].replace("films_", "").replace(".db", "")
                return datetime.strptime(name, "%Y-%m-%d_%H-%M-%S")
            except Exception:
                return datetime.min

        db_files_sorted = sorted(db_files, key=extract_date, reverse=True)

        KEEP = 20

        if len(db_files_sorted) <= KEEP:
            print("🟢 Rien à supprimer")
            return f"Backup OK → {backup_status}"

        to_delete = db_files_sorted[KEEP:]

        print("STEP 9: deleting old backups =", len(to_delete))

        for f in to_delete:
            delete_data = {
                "message": f"delete old backup {f['name']}",
                "sha": f["sha"],
                "branch": "main"
            }

            try:
                requests.delete(f["url"], json=delete_data, headers=headers, timeout=5)
                print("🗑️ OK:", f["name"])
            except Exception as e:
                print("❌ DELETE ERROR:", f["name"], e)

        print("STEP FINAL: success")
        

        return f"Backup OK → {backup_status}"

    except Exception as e:
        print("❌ ERREUR BACKUP :", e)
        return "❌ Backup crash"

    finally:
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)
            
#29B — DOWNLOAD DB
@app.route("/download_db")
def download_db():
    
    if ENV != "DEV":
        return "⛔ accès interdit"

    from flask import send_file

    if not os.path.exists(DB_PATH):
        return "❌ DB introuvable"

    now = datetime.now().strftime("%y-%m-%d_%H-%M-%S")
    filename = f"Films_{now}.db"

    return send_file(
        DB_PATH,
        as_attachment=True,
        download_name=filename
    )
    
#29C — DOWNLOAD EXCEL
@app.route("/download_excel")
def download_excel():
    
    if ENV != "DEV":
        return "⛔ accès interdit"

    import sqlite3
    from flask import send_file

    if not os.path.exists(DB_PATH):
        return "❌ DB introuvable"

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films")
    rows = cursor.fetchall()

    # 🔥 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Films"

    headers = [desc[0] for desc in cursor.description]
    ws.append(headers)

    for row in rows:
        ws.append(row)

    conn.close()

    # 🔥 horodatage
    now = datetime.now().strftime("%y-%m-%d_%H-%M-%S")
    filename = f"Films_{now}.xlsx"

    file_path = os.path.join(BASE_DIR, filename)
    wb.save(file_path)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename
    )
    
#29D — DOWNLOAD ALL (ZIP)
@app.route("/download_all")
def download_all():
    
    # 🔥 backup GitHub avant download
    try:
        backup_db()
    except Exception as e:
        print("⚠️ backup auto échoué :", e)

    import sqlite3
    from flask import send_file
    from openpyxl import Workbook
    import zipfile

    if not os.path.exists(DB_PATH):
        return "❌ DB introuvable"

    now = datetime.now().strftime("%y-%m-%d_%H-%M-%S")

    # 🔹 noms fichiers
    db_name = f"Films_{now}.db"
    excel_name = f"Films_{now}.xlsx"
    zip_name = f"Films_{now}.zip"

    db_temp = os.path.join(BASE_DIR, db_name)
    excel_temp = os.path.join(BASE_DIR, excel_name)
    zip_path = os.path.join(BASE_DIR, zip_name)

    # 🔹 copier DB
    import shutil
    shutil.copyfile(DB_PATH, db_temp)

    # 🔹 créer Excel
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films")
    rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Films"

    headers = [desc[0] for desc in cursor.description]
    ws.append(headers)

    for row in rows:
        ws.append(row)

    conn.close()

    wb.save(excel_temp)

    # 🔹 créer ZIP
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(db_temp, db_name)
        zipf.write(excel_temp, excel_name)

    # 🔹 nettoyage des fichiers temporaires
    os.remove(db_temp)
    os.remove(excel_temp)

    return send_file(
        zip_path,
        as_attachment=True,
        download_name=zip_name
    )
            
#29E — UPLOAD DB
@app.route("/upload_db", methods=["POST"])
def upload_db():

    if ENV != "DEV":
        return "⛔ accès interdit"

    from flask import request

    file = request.files.get("file")

    if not file:
        return "❌ Aucun fichier"

    if not file.filename.endswith(".db"):
        return "❌ Format invalide"

    temp_path = DB_PATH + ".upload"

    file.save(temp_path)
    
    # 🔐 vérification DB valide
    import sqlite3

    try:
        conn = sqlite3.connect(temp_path)
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM films LIMIT 1")
        conn.close(
    except:
        os.remove(temp_path)
        return "❌ DB invalide"

    # 🔥 sécurité taille minimale
    if os.path.getsize(temp_path) < 1000:
        os.remove(temp_path)
        return "❌ Fichier invalide"

    requests.get("https://ton-app.onrender.com/backup_db")

    # 🔥 remplacement sécurisé
    os.replace(temp_path, DB_PATH)

    reset_cache()

    return """
    <h2>✅ Base remplacée</h2>
    <a href="/">⬅️ Retour</a>
    """
            
#30 — HEALTH
@app.route("/health")
def health():
    return "OK", 200
    
#31 — STARTUP
@app.before_request
def startup():
    if not hasattr(app, "initialized"):
        init_app()
        app.initialized = True
        
#32 — ADMIN
@app.route("/force_restore")
def force_restore():
    print("🔥 RESTORE FORCÉ")
    restore_db()
    return "RESTORE DONE"
    
#33 — RUN

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
    

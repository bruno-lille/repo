import sqlite3
from openpyxl import load_workbook

wb = load_workbook("films.xlsx")
ws = wb.active

conn = sqlite3.connect("films.db")
cursor = conn.cursor()

for row in ws.iter_rows(min_row=2):

    cursor.execute("""
    INSERT INTO films (
        disc_id, emplacement, type, titre, allocine,
        tmdb_id, jaquette, annee, genres, resume, casting, ordre
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        str(row[4].value or ""),
        str(row[0].value or ""),
        str(row[2].value or ""),
        str(row[5].value or ""),
        str(row[6].value or ""),
        str(row[12].value or ""),
        str(row[13].value or ""),
        str(row[14].value or ""),
        str(row[15].value or ""),
        str(row[16].value or ""),
        str(row[17].value or ""),
        row[10].value
    ))

conn.commit()
conn.close()

print("✅ Import terminé")
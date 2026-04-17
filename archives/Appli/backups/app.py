# -*- coding: utf-8 -*-
DATA = None
def reset_cache():
    global DATA
    DATA = None
from flask import Flask, request, redirect
from openpyxl import load_workbook

import urllib.parse
import requests
import re

app = Flask(__name__)


API_KEY = "0458b97834e41cd8c1e0f3364d6703d6"


COL = {
    "EMPLACEMENT": 1,
    "TYPE": 3,
    "DISC_ID": 5,   # ✅ colonne E
    "TITRE": 6,
    "ALLOCINE": 7,
    "TMDB_ID": 13,
    "JAQUETTE": 14,
    "ANNEE": 15,
    "GENRES": 16,
    "RESUME": 17,
    "CASTING": 18,
    "ORDRE": 11
}

# ----------- CHARGEMENT SQL LIGHT -----------
def load_data():
    import sqlite3

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM films
    ORDER BY
        CASE WHEN ordre IS NULL THEN 999999 ELSE ordre END ASC,
        annee ASC
    """)
    rows = cursor.fetchall()

    titres, liens = [], []
    col_A, col_C = [], []
    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []
    disc_ids, ordres = [], []

    for row in rows:
        disc_ids.append(row[1])
        col_A.append(row[2])
        col_C.append(row[3])
        titres.append(row[4])
        liens.append(row[5])
        tmdb_ids.append(row[6])
        jaquettes.append(row[7])
        annees.append(row[8])
        genres.append(row[9])
        resumes.append(row[10])
        casting.append(row[11])
        ordres.append(row[12])

    conn.close()

    return titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids, ordres
    
def load_data_cached():
    global DATA

    if DATA is None:
        DATA = load_data()

    return DATA


def short_text(text, max_len=140):
    return text if len(text) <= max_len else text[:max_len] + "..."

      
        
def search_tmdb_multi(query):

    results = []

    def fetch(lang):
        try:
            url = "https://api.themoviedb.org/3/search/movie"
            params = {
                "api_key": API_KEY,
                "query": query,
                "language": lang,
                "region": "FR",
                "include_adult": False
            }

            data = requests.get(url, params=params).json()

            for film in data.get("results", []):
                poster = film.get("poster_path")
                img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""

                results.append({
                    "id": film["id"],
                    "title": film["title"],
                    "year": film.get("release_date", "")[:4],
                    "img": img
                })
        except:
            pass

    # 🔥 recherche FR
    fetch("fr-FR")

    # 🔥 fallback anglais
    if len(results) < 5:
        fetch("en-US")

    # 🔥 suppression doublons
    unique = {film["id"]: film for film in results}

    # 🔥 tri par année décroissante
    sorted_results = sorted(
        unique.values(),
        key=lambda x: (x["year"] or "0", x["title"]),
        reverse=True
    )

    return sorted_results[:5]

# ----------- EXTRACTION ID TMDB -----------

    
def extract_tmdb_id(value):

    value = value.strip()

    # cas simple : nombre direct
    if value.isdigit():
        return value

    # cas URL TMDB (robuste)
    match = re.search(r'/movie/(\d+)', value)
    if match:
        return match.group(1)

    return None
    
def generate_disc_id(ws):

    max_id = 0

    for row in ws.iter_rows(min_row=2):

        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val.startswith("DISC-"):
            try:
                num = int(val.replace("DISC-", ""))
                if num > max_id:
                    max_id = num
            except:
                pass

    new_id = max_id + 1

    return f"DISC-{str(new_id).zfill(5)}"
    
def find_row_by_disc_id(ws, disc_id):

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val == disc_id:
            return idx

    return None


# ----------- STYLE GLOBAL -----------
def get_style():
    return """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    
    
    <script>
    function updateCount() {

        let q = document.getElementById("search").value;

        if (q.length === 0) {
            document.getElementById("count").innerText = "";
            return;
        }

        fetch("/count?q=" + encodeURIComponent(q))
            .then(res => res.text())
            .then(data => {
                document.getElementById("count").innerText =
                    data + " résultats";
            });
    }
    </script>
    
    <script>
    function setTMDB(id) {
        const input = document.getElementById("tmdb_input");

        if (!input) {
            alert("Champ TMDB introuvable");
            return;
        }

        input.value = id;
        input.focus();

        window.scrollTo({ top: 0, behavior: "smooth" });
    }

    function setTitle(title) {
        const input = document.getElementById("title");

        if (!input) {
            alert("Champ titre introuvable");
            return;
        }

        input.value = title;
        input.focus();
        input.style.background = "#2a9df4";

        setTimeout(() => {
            input.style.background = "";
        }, 300);

        window.scrollTo({ top: 0, behavior: "smooth" });
    }
    </script>
    
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}

    input {
        padding:14px;
        width:85%;
        font-size:16px;
        border-radius:10px;
        border:none;
        text-align:center;
    }

    .card {
        background:#1c1c1c;
        padding:15px;
        border-radius:16px;
        margin:15px auto;
        width:95%;
        max-width:420px;
    }

    img {width:100%; border-radius:12px;}

    .titre a {
        color:#4fc3f7;
        font-size:18px;
        font-weight:bold;
        text-decoration:none;
    }

    .btn-row {
        display:flex;
        gap:8px;
        margin-top:12px;
    }

    .btn {
        flex:1;
        padding:12px;
        border-radius:10px;
        text-decoration:none;
        font-weight:bold;
        font-size:14px;
    }
    
    .btn:hover {
        opacity:0.85;
    }

    .allocine {
        background:#2a9df4;
        color:white;
        font-weight:bold;
    }
    .new {background:#e50914; color:white;}
    .update {
        background:#ffb74d;
        color:black;
    }
    .retour {background:#444; color:white;}
    </style>
    """
#-----------Appli---------------

import unicodedata

def normalize(text):

    text = text.lower()

    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')

    text = re.sub(r'[^a-z0-9]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()

    return text

# ----------- HOME -----------
@app.route("/", methods=["GET"])
def home():

    show_suggest = request.args.get("suggest")

    titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids, ordres = load_data_cached()

    query_raw = request.args.get("q", "").strip()
    mode = request.args.get("mode")

    exact_mode = False

    if query_raw.startswith("(") and query_raw.endswith(")"):
        exact_mode = True
        query = query_raw[1:-1].strip()
    else:
        query = query_raw

    html = get_style()

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>

        <form>
           
            
            <div id="count" style="font-size:20px; margin-bottom:10px;"></div>

            <input name="q" id="search" placeholder="Tape un film..." onkeyup="updateCount()">

            <div class="card">
                <div class="btn-row">
                    <button class="btn allocine" name="mode" value="fast">🔎 Recherche</button>
                    <button class="btn new" name="mode" value="poster">🎬 Jaquettes</button>
                    
                </div>
            </div>
        </form>
        """

    
        
    results = []

    for i, titre in enumerate(titres):

        q = normalize(query)
        t = normalize(titre)

        if exact_mode:

            if q == t:
                results.append((300, i))

            elif re.search(r'^' + re.escape(q) + r'\b', t):
                results.append((200, i))

            elif re.search(r'\b' + re.escape(q) + r'\b', t):
                results.append((100, i))

        else:
            if q in t:
                results.append((100, i))
                
    results.sort(key=lambda x: (
        int(ordres[x[1]]) if str(ordres[x[1]]).isdigit() else 999999,
        int(annees[x[1]]) if str(annees[x[1]]).isdigit() else 0
    ))

       

    query_encoded = urllib.parse.quote(query_raw)

    if results:
        
        bloc = f"""
         <h2>{len(results)} résultats</h2>

            <div class="card">
                <div class="btn-row">
                    <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
                    <a class="btn new" href="/">❌ Annuler</a>
                </div>
            </div>
            """
        
        #bloc = f"<h2>{len(results)} résultats</h2>"
        
        

        for score, i in results:

            poster_html = f'<img src="{jaquettes[i]}">' if jaquettes[i] else ""

            bloc += f"""
            <div class="card">
                {poster_html}

                <div class="titre">

                    <div style="color:#888; font-size:12px; margin-bottom:4px;">
                        🆔 {disc_ids[i]}
                    </div>

                    <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                        {titres[i]}
                    </a>
                </div>

                <div style="color:#b0b0b0; font-size:16px; margin-top:6px; font-weight:500;">
                    📀 {col_C[i]} &nbsp;&nbsp; 📁 {col_A[i]} &nbsp;&nbsp; 📅 {annees[i]}
                </div>

                <div>{genres[i]}</div>
                <div>{casting[i]}</div>

                <div class="btn-row">
                    <a class="btn allocine" href="{liens[i]}" target="_blank">Allociné</a>
                    <a class="btn update" href="/suggest_update/{disc_ids[i]}?q={query_encoded}">🔄 Corriger</a>
                    <a class="btn new" href="/delete/{disc_ids[i]}"
                           onclick="return confirm('Supprimer ce film ?')">
                           🗑 Supprimer
                        </a>
                </div>
            </div>
            """

        bloc += f"""
        <div class="card">
            <a class="btn allocine" href="/?q={query_encoded}&mode={mode}&suggest=1">
                🔎 Suggestions TMDB
            </a>
        </div>
        """

        if show_suggest:
            results_tmdb = search_tmdb_multi(query)

            for score, i in results:
                results_tmdb += search_tmdb_multi(titres[i])

            unique = {film["id"]: film for film in results_tmdb}

            bloc += "<h2>Suggestions TMDB</h2>"

            
                
            for film in unique.values():
                bloc += f"""
                <div class="card">
                    <a href="https://www.themoviedb.org/movie/{film['id']}" target="_blank">
                        <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
                    </a>
                    <div>{film['title']} ({film['year']})</div>

                    <div class="btn-row">
                        <a class="btn new" href="/add/{film['id']}?q={query_encoded}">
                            ➕ Ajouter
                        </a>
                    </div>
                </div>
                """

        bloc += """
        <div class="card">
            <div class="btn-row">
                <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
                <a class="btn new" href="/">❌ Annuler</a>
            </div>
        </div>
        """

        return html + bloc

   
    # ----------- AUCUN RESULTAT -----------
    results_tmdb = search_tmdb_multi(query)

    bloc = "<h2>❌ Inexistant</h2>"

    # ✅ 1. suggestions TMDB
    for film in results_tmdb:
        bloc += f"""
        <div class="card">
            <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
            <div>{film['title']} ({film['year']})</div>

            <div class="btn-row">
                <a class="btn new" href="/add/{film['id']}">➕ Ajouter</a>
            </div>
        </div>
        """

    # ✅ 2. boutons externes
    bloc += f"""
    <div class="card">
        <div class="btn-row">
            <a class="btn allocine" href="https://www.themoviedb.org/search/movie?query={query_encoded}" target="_blank">
                🔎 Voir + TMDB
            </a>
            <a class="btn allocine" href="https://www.google.com/search?q={query_encoded}+allocine+fichefilm" target="_blank">
                🎬 Voir + Allociné
            </a>
        </div>
    </div>
    """

    # ✅ 3. ajout manuel
    bloc += f"""
    <div class="card">
        <h3>✍️ Ajout manuel</h3>

        <form method="post" action="/manual_add">

            <input name="title" value="{query}" required><br><br>


            <input name="emplacement" placeholder="📁 Emplacement" required><br><br>

            <input name="type" placeholder="📀 Type (DVD/BLURAY)" required><br><br>

           <input name="allocine" placeholder="🔗 Lien Allociné"><br><br>

            <input name="tmdb_input" placeholder="ID ou lien TMDB"><br><br>

            <button class="btn new">➕ Ajouter</button>

        </form>
    </div>
    """

    return html + bloc


# ----------- PAGE CORRECTION -----------
@app.route("/suggest_update/<disc_id>")
def suggest_update(disc_id):

    import sqlite3
    import urllib.parse

    query = request.args.get("q", "")
    query_encoded = urllib.parse.quote(query)
    results_tmdb = search_tmdb_multi(query)

    conn = sqlite3.connect("films.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()
    conn.close()

    if not film:
        return "❌ Film introuvable"

    title = film["titre"] or ""
    emplacement = film["emplacement"] or ""
    type_disc = film["type"] or ""
    allocine = film["allocine"] or ""
    tmdb_id = film["tmdb_id"] or ""
    ordre = film["ordre"] or ""

    html = get_style() + f"""
    <h2>🛠 Modification / Mise à jour</h2>

    <div class="card">
        <a class="btn retour" href="/?q={query}">⬅️ Retour</a>
        <a class="btn new" href="/">❌ Annuler</a>
    </div>

    <div class="card">
        <h3>✏️ Modifier complètement</h3>

        <form action="/manual_update/{disc_id}?q={query_encoded}" method="post">
    
            <input id="title" name="title" value="{title}" placeholder="Titre"><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement"><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)"><br><br>

            <input name="allocine" value="{allocine}" placeholder="🔗 Allociné"><br><br>

            <input name="ordre" value="{ordre}" placeholder="🔢 Ordre" type="number" min="1"><br><br>

            <input id="tmdb_input" name="tmdb_input" value="{tmdb_id}" placeholder="ID ou lien TMDB"><br><br>

            <button type="submit" class="btn update">💾 Mettre à jour</button>

        </form>
            </div>
            """

    for film in results_tmdb:

        title_safe = film["title"].replace("'", "\\'")

        html += f"""
        <div class="card">
            <a href="https://www.themoviedb.org/movie/{film['id']}" target="_blank">
                <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
            </a>

            <div class="titre">{film['title']} ({film['year']})</div>

            <div class="btn-row">
                <button type="button" class="btn update"
                    onclick="window.location.href='/confirm_update/{disc_id}/{film['id']}?q={query_encoded}'">
                    ⚡ Remplir automatiquement
                </button>
            </div>
        </div>
        """

    return html


# ----------- UPDATE AUTO -----------
@app.route("/confirm_update/<disc_id>/<int:tmdb_id>")
def confirm_update(disc_id, tmdb_id, query=""):

    import sqlite3

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {"api_key": API_KEY, "language": "fr-FR", "append_to_response": "credits"}

    data = requests.get(url, params=params).json()

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    cursor.execute("""
    UPDATE films SET
        tmdb_id = ?,
        jaquette = ?,
        annee = ?,
        genres = ?,
        resume = ?,
        casting = ?
    WHERE disc_id = ?
    """, (
        tmdb_id,
        f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}",
        data.get("release_date", "")[:4],
        ", ".join([g["name"] for g in data.get("genres", [])]),
        data.get("overview", ""),
        ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]]),
        disc_id
    ))

    conn.commit()
    conn.close()

    reset_cache()

    query = request.args.get("q", "")
    return redirect(f"/?q={query}")



# ----------- UPDATE MANUEL (ID OU URL) -----------
@app.route("/manual_update/<disc_id>", methods=["POST"])
def manual_update(disc_id):

    import sqlite3

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    # 🔥 vérifie si le film existe
    cursor.execute("SELECT id FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()

    if not film:
        conn.close()
        return "❌ Introuvable"

    # 🔥 mise à jour des champs
    cursor.execute("""
        UPDATE films SET
            titre = ?,
            emplacement = ?,
            type = ?,
            allocine = ?,
            ordre = ?,
            tmdb_id = ?,
            jaquette = ?,
            annee = ?,
            genres = ?,
            resume = ?,
            casting = ?
        WHERE disc_id = ?
    """, (
        request.form.get("title", ""),
        request.form.get("emplacement", ""),
        request.form.get("type", ""),
        request.form.get("allocine", ""),
        int(request.form.get("ordre")) if request.form.get("ordre") else None,
        request.form.get("tmdb_input", ""),
        request.form.get("poster", ""),
        request.form.get("year", ""),
        request.form.get("genres", ""),
        request.form.get("overview", ""),
        request.form.get("cast", ""),
        disc_id
    ))

    conn.commit()
    conn.close()
    reset_cache()

    # 🔥 TMDB optionnel
    raw = request.form.get("tmdb_input", "")
    tmdb_id = extract_tmdb_id(raw) if raw else None

    if tmdb_id:
        query = request.args.get("q", "")
        return confirm_update(disc_id, int(tmdb_id), query)
        
        reset_cache()

    query = request.args.get("q", "")
    return redirect(f"/?q={query}")


# ----------- PAGE AJOUT FILM -----------
@app.route("/add/<int:tmdb_id>")
def add_movie(tmdb_id):

    emplacement = request.args.get("emplacement", "")
    type_disc = request.args.get("type", "")
    allocine = request.args.get("allocine", "")

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params).json()

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]
    genres = ", ".join([g["name"] for g in data.get("genres", [])])
    overview = data.get("overview", "")
    poster = f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}"
    cast = ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]])
    query_allocine = urllib.parse.quote(f"{title} {year} allocine fichefilm")

    return get_style() + f"""
    <h2>➕ Ajouter ce film</h2>

    <div class="card">
        <img src="{poster}">
        <h3>{title} ({year})</h3>
        <p>{genres}</p>
        <p>{cast}</p>
        <p>{overview}</p>
    </div>

    <div class="card">
        <form method="post" action="/confirm_add">

            <input type="hidden" name="tmdb_id" value="{tmdb_id}">
            <input type="hidden" name="title" value="{title}">
            <input type="hidden" name="year" value="{year}">
            <input type="hidden" name="genres" value="{genres}">
            <input type="hidden" name="overview" value="{overview}">
            <input type="hidden" name="cast" value="{cast}">
            <input type="hidden" name="poster" value="{poster}">
            
            <input name="title" value="{title}" required><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement" required><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)" required><br><br>



            <input name="allocine" 
                  value="https://www.google.com/search?q={query_allocine}">

            <div class="btn-row">
                <a class="btn allocine" 
                   href="https://www.google.com/search?q={query_allocine}" 
                   target="_blank">
                    🎬 Chercher Allociné
                </a>
            </div>
            <br>

            <button class="btn new">✅ Ajouter</button>
        </form>
    </div>

    <div class="card">
        <a class="btn retour" href="/">⬅️ Retour</a>
    </div>
    """
            

# ----------- CONFIRM ADD -----------
@app.route("/confirm_add", methods=["POST"])
def confirm_add():

    import sqlite3

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    # 🔥 générer DISC_ID automatiquement
    cursor.execute("SELECT disc_id FROM films ORDER BY id DESC LIMIT 1")
    last = cursor.fetchone()

    if last and last[0].startswith("DISC-"):
        num = int(last[0].replace("DISC-", ""))
        new_id = f"DISC-{str(num + 1).zfill(5)}"
    else:
        new_id = "DISC-00001"

    cursor.execute("""
    INSERT INTO films (
        disc_id, emplacement, type, titre, allocine,
        tmdb_id, jaquette, annee, genres, resume, casting, ordre
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        new_id,
        request.form.get("emplacement", ""),
        request.form.get("type", ""),
        request.form.get("title", ""),
        request.form.get("allocine", ""),
        request.form.get("tmdb_id", ""),
        request.form.get("poster", ""),
        request.form.get("year", ""),
        request.form.get("genres", ""),
        request.form.get("overview", ""),
        request.form.get("cast", ""),
        None
    ))

    conn.commit()
    conn.close()

    reset_cache()

    return get_style() + """
    <h2>✅ Film ajouté !</h2>

    <div class="card">
        <a class="btn allocine" href="/">🏠 Retour accueil</a>
    </div>
    """
    
# ----------- suppression ligne -----------

@app.route("/delete/<disc_id>")
def delete_movie(disc_id):

    import sqlite3

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM films WHERE disc_id = ?", (disc_id,))

    conn.commit()
    conn.close()

    reset_cache()

    return "<h2>🗑 Film supprimé</h2><a href='/'>Retour</a>"
#------------COMPTEUR--------------


@app.route("/count")
def count():

    import sqlite3

    query_raw = request.args.get("q", "").strip()

    if not query_raw:
        return "0"

    if query_raw.startswith("(") and query_raw.endswith(")"):
        query = query_raw[1:-1].strip()
    else:
        query = query_raw

    conn = sqlite3.connect("films.db")
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM films WHERE LOWER(titre) LIKE ?", (f"%{query.lower()}%",))
    result = cursor.fetchone()[0]

    conn.close()

    return str(result)
    
    #----------Tout Automatique------------
    

@app.route("/prefill/<disc_id>/<int:tmdb_id>")
def prefill(disc_id, tmdb_id):

    import sqlite3

    # 🔥 récupération TMDB
    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params).json()

    poster_path = data.get("poster_path")
    poster = f"https://image.tmdb.org/t/p/w300{poster_path}" if poster_path else ""

    genres = ", ".join([g["name"] for g in data.get("genres", [])]) if data.get("genres") else ""
    overview = data.get("overview", "")
    cast = ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]])

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]

    # 🔥 récupération données existantes (SQLite)
    conn = sqlite3.connect("films.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM films WHERE disc_id = ?", (disc_id,))
    film = cursor.fetchone()

    conn.close()

    emplacement = film["emplacement"] if film else ""
    type_disc = film["type"] if film else ""
    allocine = film["allocine"] if film else ""
    ordre = film["ordre"] if film else ""

    return get_style() + f"""
    <h2>✨ Pré-remplissage</h2>

    <div class="card">
        <h3>{title} ({year})</h3>
    </div>

    <div class="card">
        <form action="/manual_update/{disc_id}" method="post">
        
            

            <input name="title" value="{title}" placeholder="Titre"><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement"><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)"><br><br>

            <input name="allocine" value="{allocine}" placeholder="🔗 Allociné"><br><br>

            <input name="ordre" value="{ordre}" placeholder="🔢 Ordre" type="number" min="1"><br><br>

            <input id="tmdb_input" name="tmdb_input" value="{tmdb_id}" placeholder="ID ou lien TMDB">

            <button class="btn new">✅ Valider</button>

            <div class="card">
                <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
            </div>

        </form>
    </div>
    """
#-------------SAUVEGARDE VERS GITHUB----------------

@app.route("/backup_db", methods=["GET", "HEAD"])
def backup_db():

    import os
    import base64
    import requests
    import sqlite3
    import shutil
    from datetime import datetime

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DB_PATH = os.path.join(BASE_DIR, "films.db")
    TMP_PATH = os.path.join(BASE_DIR, "backup_temp.db")

    token = os.getenv("GITHUB_TOKEN")
    repo = os.getenv("GITHUB_REPO")

    headers = {
        "Authorization": f"token {token}"
    }

    # 🔍 Vérification DB
    if not os.path.exists(DB_PATH):
        return "❌ films.db introuvable"

    try:
        # 🔥 1. Forcer écriture SQLite (flush WAL)
        conn = sqlite3.connect(DB_PATH)
        conn.commit()
        conn.execute("PRAGMA wal_checkpoint(FULL);")
        conn.close()

        # 🔥 2. Copier une version stable
        shutil.copyfile(DB_PATH, TMP_PATH)

        # 🔥 3. Lire la copie (jamais l’original)
        with open(TMP_PATH, "rb") as f:
            content = base64.b64encode(f.read()).decode()

        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"backups/films_{now}.db"

        url = f"https://api.github.com/repos/{repo}/contents/{filename}"

        data = {
            "message": f"backup {now}",
            "content": content,
            "branch": "main"
        }

        r = requests.put(url, json=data, headers=headers)

        return f"✅ Backup OK → {r.status_code}"

    except Exception as e:
        return f"❌ Erreur backup : {str(e)}"

    finally:
        # 🧹 Nettoyage
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)
    
#if __name__ == "__main__":
#    app.run(host="0.0.0.0", port=5000)
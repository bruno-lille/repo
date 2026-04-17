from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import urllib.parse
import requests

app = Flask(__name__)

API_KEY = "YOUR_TMDB_API_KEY"

# ----------- CHARGEMENT EXCEL -----------
try:
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens = [], []
    col_A, col_C = [], []

    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []

    for row in ws.iter_rows(min_row=2):

        titre = str(row[5].value or "").strip()

        if row[6].hyperlink:
            lien = row[6].hyperlink.target
        else:
            lien = str(row[6].value or "")

        col_A.append(str(row[0].value or ""))
        col_C.append(str(row[2].value or ""))

        tmdb_ids.append(str(row[12].value or ""))
        jaquettes.append(str(row[13].value or ""))
        annees.append(str(row[14].value or ""))
        genres.append(str(row[15].value or ""))
        resumes.append(str(row[16].value or ""))
        casting.append(str(row[17].value or ""))

        titres.append(titre)
        liens.append(lien)

except Exception as e:
    titres, liens, col_A, col_C = [], [], [], []
    tmdb_ids, jaquettes, annees, genres, resumes, casting = [], [], [], [], [], []
    erreur = str(e)


def short_text(text, max_len=140):
    text = str(text)
    return text if len(text) <= max_len else text[:max_len] + "..."


def get_tmdb_data(query):
    try:
        url = "https://api.themoviedb.org/3/search/movie"
        params = {"api_key": API_KEY, "query": query, "language": "fr-FR"}
        r = requests.get(url, params=params)
        data = r.json()

        if data["results"]:
            film = data["results"][0]
            poster = film.get("poster_path")
            tmdb_id = film.get("id")

            if poster:
                img = f"https://image.tmdb.org/t/p/w600_and_h900_face{poster}"
            else:
                img = "https://image.tmdb.org/t/p/w600_and_h900_face/5jl5sfhTeAlp60rM9GxoDA2dmX9.jpg"

            return img, tmdb_id

    except:
        pass

    # fallback ET
    return "https://image.tmdb.org/t/p/w600_and_h900_face/5jl5sfhTeAlp60rM9GxoDA2dmX9.jpg", ""


@app.route("/", methods=["GET"])
def home():
    query = request.args.get("q", "")
    mode = request.args.get("mode")

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}
    input {padding:15px; width:85%; font-size:18px; border-radius:12px; border:none;}
    button {padding:12px 20px; font-size:16px; border-radius:10px; border:none; margin:5px;}
    .fast {background:#444; color:white;}
    .poster {background:#e50914; color:white;}
    .card {background:#1c1c1c; padding:15px; border-radius:16px; margin:15px auto; width:95%; max-width:420px;}
    img {width:100%; border-radius:12px; margin-bottom:10px;}
    .titre {font-size:20px; font-weight:bold;}
    .meta {color:#aaa; font-size:14px;}
    .info {color:#ddd; font-size:14px; text-align:left;}
    .resume {color:#bbb; font-size:13px; text-align:left;}
    .btn-row {display:flex; gap:8px; margin-top:12px;}
    .btn {flex:1; padding:10px; border-radius:10px; text-decoration:none; font-size:14px;}
    .allocine {background:#00c3ff; color:black;}
    .retour {background:#444; color:white;}
    .new {background:#e50914; color:white;}
    </style>
    """

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film..."><br>
            <button class="fast" name="mode" value="fast">🔎 Rapide</button>
            <button class="poster" name="mode" value="poster">🎬 Jaquettes</button>
        </form>
        """

    query_clean = query.lower().strip()
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        fuzzy_score = fuzz.partial_ratio(query_clean, titre_clean)
        if fuzzy_score > 85:
            score += fuzzy_score

        if score >= 100:
            results.append((score, i))

    results.sort(key=lambda x: (titres[x[1]].lower(), annees[x[1]]))

    query_encoded = urllib.parse.quote(query)

    if results:
        bloc = f"<h2>{len(results)} résultats</h2>"

        for score, i in results:

            poster_html = ""
            if mode == "poster" and jaquettes[i] not in ["", "nan"]:
                poster_html = f'''
                <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                    <img src="{jaquettes[i]}">
                </a>
                '''

            bloc += f"""
            <div class="card">
                {poster_html}
                <div class="titre">{titres[i]}</div>
                <div class="meta">📀 {col_C[i]} • 📁 {col_A[i]} • 📅 {annees[i]}</div>
                <div class="info">🎭 {genres[i]}</div>
                <div class="info">⭐ {casting[i]}</div>
                <div class="resume">{short_text(resumes[i])}</div>

                <div class="btn-row">
                    <a class="btn retour" href="javascript:history.back()">⬅️</a>
                    <a class="btn allocine" href="{liens[i]}" target="_blank">🎬 Allociné</a>
                    <a class="btn new" href="/">🔄</a>
                </div>
            </div>
            """

        bloc += f"""
        <div class="card">
            <div class="btn-row">
                <a class="btn allocine" href="https://www.themoviedb.org/search/movie?query={query_encoded}" target="_blank">🔎 Autres TMDB</a>
                <a class="btn allocine" href="https://www.google.com/search?q={query_encoded}+allocine+fichefilm" target="_blank">🎬 Autres Allociné</a>
            </div>
        </div>
        """

        return html + bloc

    # ----------- AUCUN RESULTAT -----------
    img, tmdb_id = get_tmdb_data(query)

    tmdb_link = f"https://www.themoviedb.org/movie/{tmdb_id}" if tmdb_id else f"https://www.themoviedb.org/search/movie?query={query_encoded}"
    allocine_search = f"https://www.google.com/search?q={query_encoded}+allocine+fichefilm"

    return html + f"""
    <h2>❌ Inexistant dans la liste</h2>
    <p>🔎 Voir d'autres résultats :</p>

    <div class="card">
        <a href="{tmdb_link}" target="_blank">
            <img src="{img}">
        </a>

        <div class="titre">{query}</div>

        <div class="btn-row">
            <a class="btn retour" href="javascript:history.back()">⬅️</a>
            <a class="btn allocine" href="{allocine_search}" target="_blank">🎬 Allociné</a>
            <a class="btn new" href="/">🔄</a>
        </div>
    </div>
    """


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
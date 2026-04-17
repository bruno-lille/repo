# -*- coding: utf-8 -*-
from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import urllib.parse
import requests
import re

app = Flask(__name__)

API_KEY = "0458b97834e41cd8c1e0f3364d6703d6"


COL = {
    "EMPLACEMENT": 1,
    "TYPE": 3,
    "DISC_ID": 5,   # ✅ colonne E
    "TITRE": 6,
    "ALLOCINE": 7,
    "TMDB_ID": 13,
    "JAQUETTE": 14,
    "ANNEE": 15,
    "GENRES": 16,
    "RESUME": 17,
    "CASTING": 18
}

# ----------- CHARGEMENT EXCEL -----------
def load_data():
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens = [], []
    col_A, col_C = [], []
    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []
    
    disc_ids = []

    for row in ws.iter_rows(min_row=2):

        titre = str(row[COL["TITRE"] - 1].value or "").strip()

        if row[6].hyperlink:
            lien = row[6].hyperlink.target
        else:
            lien = str(row[6].value or "")

        col_A.append(str(row[0].value or ""))
        col_C.append(str(row[2].value or ""))

        tmdb_ids.append(str(row[12].value or ""))
        jaquettes.append(str(row[13].value or ""))
        annees.append(str(row[14].value or ""))
        genres.append(str(row[15].value or ""))
        resumes.append(str(row[16].value or ""))
        casting.append(str(row[17].value or ""))

        titres.append(titre)
        liens.append(lien)
        
        disc_ids.append(str(row[COL["DISC_ID"] - 1].value or ""))

    return titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids


def short_text(text, max_len=140):
    return text if len(text) <= max_len else text[:max_len] + "..."


# ----------- TMDB SEARCH -----------
# def search_tmdb_multi(query):
    # try:
        # url = "https://api.themoviedb.org/3/search/movie"
        # params = {
            # "api_key": API_KEY,
            # "query": query,
            # "language": "fr-FR",
            # "region": "FR",
            # "include_adult": False
        # }
        # data = requests.get(url, params=params).json()

        # results = []
        # for film in data.get("results", [])[:5]:
            # poster = film.get("poster_path")
            # img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""
            # results.append({
                # "id": film["id"],
                # "title": film["title"],
                # "year": film.get("release_date", "")[:4],
                # "img": img
            # })
        # return results
    # except:
        # return []
        
        
def search_tmdb_multi(query):

    results = []

    def fetch(lang):
        try:
            url = "https://api.themoviedb.org/3/search/movie"
            params = {
                "api_key": API_KEY,
                "query": query,
                "language": lang,
                "region": "FR",
                "include_adult": False
            }

            data = requests.get(url, params=params).json()

            for film in data.get("results", []):
                poster = film.get("poster_path")
                img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""

                results.append({
                    "id": film["id"],
                    "title": film["title"],
                    "year": film.get("release_date", "")[:4],
                    "img": img
                })
        except:
            pass

    # 🔥 recherche FR
    fetch("fr-FR")

    # 🔥 fallback anglais
    if len(results) < 5:
        fetch("en-US")

    # 🔥 suppression doublons
    unique = {film["id"]: film for film in results}

    # 🔥 tri par année décroissante
    sorted_results = sorted(
        unique.values(),
        key=lambda x: (x["year"] or "0", x["title"]),
        reverse=True
    )

    return sorted_results[:5]

# ----------- EXTRACTION ID TMDB -----------

    
def extract_tmdb_id(value):

    value = value.strip()

    # cas simple : nombre direct
    if value.isdigit():
        return value

    # cas URL TMDB (robuste)
    match = re.search(r'/movie/(\d+)', value)
    if match:
        return match.group(1)

    return None
    
def generate_disc_id(ws):

    max_id = 0

    for row in ws.iter_rows(min_row=2):

        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val.startswith("DISC-"):
            try:
                num = int(val.replace("DISC-", ""))
                if num > max_id:
                    max_id = num
            except:
                pass

    new_id = max_id + 1

    return f"DISC-{str(new_id).zfill(5)}"
    
def find_row_by_disc_id(ws, disc_id):

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        val = str(row[COL["DISC_ID"] - 1].value or "").strip()

        if val == disc_id:
            return idx

    return None


# ----------- STYLE GLOBAL -----------
def get_style():
    return """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}

    input {
        padding:14px;
        width:85%;
        font-size:16px;
        border-radius:10px;
        border:none;
        text-align:center;
    }

    .card {
        background:#1c1c1c;
        padding:15px;
        border-radius:16px;
        margin:15px auto;
        width:95%;
        max-width:420px;
    }

    img {width:100%; border-radius:12px;}

    .titre a {
        color:#4fc3f7;
        font-size:18px;
        font-weight:bold;
        text-decoration:none;
    }

    .btn-row {
        display:flex;
        gap:8px;
        margin-top:12px;
    }

    .btn {
        flex:1;
        padding:12px;
        border-radius:10px;
        text-decoration:none;
        font-weight:bold;
        font-size:14px;
    }
    
    .btn:hover {
        opacity:0.85;
    }

    .allocine {
        background:#2a9df4;
        color:white;
        font-weight:bold;
    }
    .new {background:#e50914; color:white;}
    .update {
        background:#ffb74d;
        color:black;
    }
    .retour {background:#444; color:white;}
    </style>
    """


# ----------- HOME -----------
@app.route("/", methods=["GET"])
def home():

    show_suggest = request.args.get("suggest")

    titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting, disc_ids = load_data()

    query = request.args.get("q", "")
    mode = request.args.get("mode")

    html = get_style()

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>

        <form>
            <input name="q" placeholder="Tape un film..."><br><br>

            <div class="card">
                <div class="btn-row">
                    <button class="btn allocine" name="mode" value="fast">🔎 Recherche</button>
                    <button class="btn new" name="mode" value="poster">🎬 Jaquettes</button>
                    
                </div>
            </div>
        </form>
        """

    query_clean = query.lower().strip()
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        if fuzz.partial_ratio(query_clean, titre_clean) > 90:
            score += 100

        if score >= 100:
            results.append((score, i))

    query_encoded = urllib.parse.quote(query)

    if results:
        
        bloc = f"""
         <h2>{len(results)} résultats</h2>

            <div class="card">
                <div class="btn-row">
                    <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
                    <a class="btn new" href="/">❌ Annuler</a>
                </div>
            </div>
            """
        
        #bloc = f"<h2>{len(results)} résultats</h2>"
        
        

        for score, i in results:

            poster_html = f'<img src="{jaquettes[i]}">' if jaquettes[i] else ""

            bloc += f"""
            <div class="card">
                {poster_html}

                <div class="titre">

                    <div style="color:#888; font-size:12px; margin-bottom:4px;">
                        🆔 {disc_ids[i]}
                    </div>

                    <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                        {titres[i]}
                    </a>
                </div>

                <div style="color:#b0b0b0; font-size:16px; margin-top:6px; font-weight:500;">
                    📀 {col_C[i]} &nbsp;&nbsp; 📁 {col_A[i]} &nbsp;&nbsp; 📅 {annees[i]}
                </div>

                <div>{genres[i]}</div>
                <div>{casting[i]}</div>

                <div class="btn-row">
                    <a class="btn allocine" href="{liens[i]}" target="_blank">Allociné</a>
                    <a class="btn update" href="/suggest_update/{disc_ids[i]}?q={query_encoded}">🔄 Corriger</a>
                    <a class="btn new" href="/delete/{disc_ids[i]}"
                           onclick="return confirm('Supprimer ce film ?')">
                           🗑 Supprimer
                        </a>
                </div>
            </div>
            """

        bloc += f"""
        <div class="card">
            <a class="btn allocine" href="/?q={query_encoded}&mode={mode}&suggest=1">
                🔎 Suggestions TMDB
            </a>
        </div>
        """

        if show_suggest:
            results_tmdb = search_tmdb_multi(query)

            for score, i in results:
                results_tmdb += search_tmdb_multi(titres[i])

            unique = {film["id"]: film for film in results_tmdb}

            bloc += "<h2>Suggestions TMDB</h2>"

            
                
            for film in unique.values():
                bloc += f"""
                <div class="card">
                    <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
                    <div>{film['title']} ({film['year']})</div>

                    <div class="btn-row">
                        <a class="btn new" href="/add/{film['id']}?q={query_encoded}">
                            ➕ Ajouter
                        </a>
                    </div>
                </div>
                """

        bloc += """
        <div class="card">
            <div class="btn-row">
                <a class="btn retour" href="javascript:history.back()">⬅️ Retour</a>
                <a class="btn new" href="/">❌ Annuler</a>
            </div>
        </div>
        """

        return html + bloc

   
    # ----------- AUCUN RESULTAT -----------
    results_tmdb = search_tmdb_multi(query)

    bloc = "<h2>❌ Inexistant</h2>"

    # ✅ 1. suggestions TMDB
    for film in results_tmdb:
        bloc += f"""
        <div class="card">
            <img src="{film['img'] or 'https://via.placeholder.com/300x450?text=No+Image'}">
            <div>{film['title']} ({film['year']})</div>

            <div class="btn-row">
                <a class="btn new" href="/add/{film['id']}">➕ Ajouter</a>
            </div>
        </div>
        """

    # ✅ 2. boutons externes
    bloc += f"""
    <div class="card">
        <div class="btn-row">
            <a class="btn allocine" href="https://www.themoviedb.org/search/movie?query={query_encoded}" target="_blank">
                🔎 Voir + TMDB
            </a>
            <a class="btn allocine" href="https://www.google.com/search?q={query_encoded}+allocine+fichefilm" target="_blank">
                🎬 Voir + Allociné
            </a>
        </div>
    </div>
    """

    # ✅ 3. ajout manuel
    bloc += f"""
    <div class="card">
        <h3>✍️ Ajout manuel</h3>

        <form method="post" action="/manual_add">

            <input name="title" value="{query}" required><br><br>


            <input name="emplacement" placeholder="📁 Emplacement" required><br><br>

            <input name="type" placeholder="📀 Type (DVD/BLURAY)" required><br><br>

           <input name="allocine" placeholder="🔗 Lien Allociné"><br><br>

            <input name="tmdb_input" placeholder="ID ou lien TMDB"><br><br>

            <button class="btn new">➕ Ajouter</button>

        </form>
    </div>
    """

    return html + bloc


# ----------- PAGE CORRECTION -----------
@app.route("/suggest_update/<disc_id>")
def suggest_update(disc_id):

    query = request.args.get("q", "")
    results_tmdb = search_tmdb_multi(query)

    wb = load_workbook("films.xlsx")
    ws = wb.active

    row = find_row_by_disc_id(ws, disc_id)

    title = ws.cell(row=row, column=COL["TITRE"]).value or ""
    emplacement = ws.cell(row=row, column=COL["EMPLACEMENT"]).value or ""
    type_disc = ws.cell(row=row, column=COL["TYPE"]).value or ""
    allocine = ws.cell(row=row, column=COL["ALLOCINE"]).value or ""

    html = get_style() + f"""
    <h2>Choisir le bon film</h2>

    <div class="card">
        <a class="btn retour" href="/?q={query}">⬅️ Retour</a>
        <a class="btn new" href="/">❌ Annuler</a>
    </div>

    <div class="card">
        <h3>✏️ Modifier complètement</h3>

        <form action="/manual_update/{disc_id}" method="post">

            <input name="title" value="{title}" placeholder="Titre">

            <input name="emplacement" value="{emplacement}"><br><br>

            <input name="type" value="{type_disc}"><br><br>

            <input name="allocine" value="{allocine}"><br><br>

            <input name="tmdb_input" placeholder="ID ou lien TMDB"><br><br>
            

            <button class="btn update">💾 Mettre à jour</button>

        </form>
    </div>
    """

    for film in results_tmdb:
        html += f"""
        <div class="card">
            <img src="{film['img']}">
            <div class="titre">{film['title']} ({film['year']})</div>
            <a class="btn update" href="/confirm_update/{disc_id}/{film['id']}">
                🔄 Mettre à jour TMDB
            </a>
        </div>
        """

    return html


# ----------- UPDATE AUTO -----------
@app.route("/confirm_update/<disc_id>/<int:tmdb_id>")
def confirm_update(disc_id, tmdb_id):

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {"api_key": API_KEY, "language": "fr-FR", "append_to_response": "credits"}

    data = requests.get(url, params=params).json()

    wb = load_workbook("films.xlsx")
    ws = wb.active

    # 🔥 recherche par DISC_ID
    row_excel = find_row_by_disc_id(ws, disc_id)

    if not row_excel:
        return "❌ DISC_ID introuvable"

    ws.cell(row=row_excel, column=COL["TMDB_ID"], value=tmdb_id)
    ws.cell(row=row_excel, column=COL["JAQUETTE"], value=f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}")
    ws.cell(row=row_excel, column=COL["ANNEE"], value=data.get("release_date", "")[:4])
    ws.cell(row=row_excel, column=COL["GENRES"], value=", ".join([g["name"] for g in data.get("genres", [])]))
    ws.cell(row=row_excel, column=COL["RESUME"], value=data.get("overview", ""))
    ws.cell(row=row_excel, column=COL["CASTING"], value=", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]]))

    wb.save("films.xlsx")

    return "<h2>✅ Mise à jour OK</h2><a href='/'>Retour</a>"



# ----------- UPDATE MANUEL (ID OU URL) -----------
@app.route("/manual_update/<disc_id>", methods=["POST"])
def manual_update(disc_id):

    wb = load_workbook("films.xlsx")
    ws = wb.active

    row = find_row_by_disc_id(ws, disc_id)

    if not row:
        return "❌ Introuvable"

    # 🔥 champs utilisateur
    ws.cell(row=row, column=COL["TITRE"], value=request.form.get("title", ""))
    ws.cell(row=row, column=COL["EMPLACEMENT"], value=request.form.get("emplacement", ""))
    ws.cell(row=row, column=COL["TYPE"], value=request.form.get("type", ""))
    ws.cell(row=row, column=COL["ALLOCINE"], value=request.form.get("allocine", ""))

    # 🔥 TMDB optionnel
    raw = request.form.get("tmdb_input", "")
    tmdb_id = extract_tmdb_id(raw) if raw else None

    if tmdb_id:
        return confirm_update(disc_id, int(tmdb_id))

    wb.save("films.xlsx")

    return "<h2>✅ Mise à jour complète</h2><a href='/'>Retour</a>"


# ----------- PAGE AJOUT FILM -----------
@app.route("/add/<int:tmdb_id>")
def add_movie(tmdb_id):

    emplacement = request.args.get("emplacement", "")
    type_disc = request.args.get("type", "")
    allocine = request.args.get("allocine", "")

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params).json()

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]
    genres = ", ".join([g["name"] for g in data.get("genres", [])])
    overview = data.get("overview", "")
    poster = f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}"
    cast = ", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]])
    query_allocine = urllib.parse.quote(f"{title} {year} allocine fichefilm")

    return get_style() + f"""
    <h2>➕ Ajouter ce film</h2>

    <div class="card">
        <img src="{poster}">
        <h3>{title} ({year})</h3>
        <p>{genres}</p>
        <p>{cast}</p>
        <p>{overview}</p>
    </div>

    <div class="card">
        <form method="post" action="/confirm_add">

            <input type="hidden" name="tmdb_id" value="{tmdb_id}">
            <input type="hidden" name="title" value="{title}">
            <input type="hidden" name="year" value="{year}">
            <input type="hidden" name="genres" value="{genres}">
            <input type="hidden" name="overview" value="{overview}">
            <input type="hidden" name="cast" value="{cast}">
            <input type="hidden" name="poster" value="{poster}">
            
            <input name="title" value="{title}" required><br><br>

            <input name="emplacement" value="{emplacement}" placeholder="📁 Emplacement" required><br><br>

            <input name="type" value="{type_disc}" placeholder="📀 Type (DVD/BLURAY)" required><br><br>



            <input name="allocine" 
                  value="https://www.google.com/search?q={query_allocine}">

            <div class="btn-row">
                <a class="btn allocine" 
                   href="https://www.google.com/search?q={query_allocine}" 
                   target="_blank">
                    🎬 Chercher Allociné
                </a>
            </div>
            <br>

            <button class="btn new">✅ Ajouter</button>
        </form>
    </div>

    <div class="card">
        <a class="btn retour" href="/">⬅️ Retour</a>
    </div>
    """
            

# ----------- CONFIRM ADD -----------
@app.route("/confirm_add", methods=["POST"])
def confirm_add():

    wb = load_workbook("films.xlsx")
    ws = wb.active

    # 🔥 génération DISC_ID
    disc_id = generate_disc_id(ws)

    row = ws.max_row + 1

    ws.cell(row=row, column=COL["EMPLACEMENT"], value=request.form["emplacement"])
    ws.cell(row=row, column=COL["TYPE"], value=request.form["type"])
    ws.cell(row=row, column=COL["DISC_ID"], value=disc_id)  # ✅ clé
    ws.cell(row=row, column=COL["TITRE"], value=request.form["title"])
    ws.cell(row=row, column=COL["ALLOCINE"], value=request.form["allocine"])

    ws.cell(row=row, column=COL["TMDB_ID"], value=request.form["tmdb_id"])
    ws.cell(row=row, column=COL["JAQUETTE"], value=request.form["poster"])
    ws.cell(row=row, column=COL["ANNEE"], value=request.form["year"])
    ws.cell(row=row, column=COL["GENRES"], value=request.form["genres"])
    ws.cell(row=row, column=COL["RESUME"], value=request.form["overview"])
    ws.cell(row=row, column=COL["CASTING"], value=request.form["cast"])

    wb.save("films.xlsx")

    return get_style() + """
    <h2>✅ Film ajouté !</h2>

    <div class="card">
        <a class="btn allocine" href="/">🏠 Retour accueil</a>
    </div>
    """
    
# ----------- suppression ligne -----------

@app.route("/delete/<disc_id>")
def delete_movie(disc_id):

    wb = load_workbook("films.xlsx")
    ws = wb.active

    row = find_row_by_disc_id(ws, disc_id)

    if row:
        ws.delete_rows(row)
        wb.save("films.xlsx")
        return "<h2>🗑 Film supprimé</h2><a href='/'>Retour</a>"

    return "<h2>❌ Film introuvable</h2><a href='/'>Retour</a>"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
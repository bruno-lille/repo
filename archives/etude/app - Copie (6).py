from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import urllib.parse
import requests

app = Flask(__name__)

API_KEY = "0458b97834e41cd8c1e0f3364d6703d6"

COL = {
    "EMPLACEMENT": 1,
    "TYPE": 3,
    "TITRE": 6,
    "ALLOCINE": 7,
    "TMDB_ID": 13,
    "JAQUETTE": 14,
    "ANNEE": 15,
    "GENRES": 16,
    "RESUME": 17,
    "CASTING": 18
}

# ----------- CHARGEMENT EXCEL -----------
def load_data():
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens = [], []
    col_A, col_C = [], []
    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []

    for row in ws.iter_rows(min_row=2):

        titre = str(row[COL["TITRE"] - 1].value or "").strip()

        if row[6].hyperlink:
            lien = row[6].hyperlink.target
        else:
            lien = str(row[6].value or "")

        col_A.append(str(row[0].value or ""))
        col_C.append(str(row[2].value or ""))

        tmdb_ids.append(str(row[12].value or ""))
        jaquettes.append(str(row[13].value or ""))
        annees.append(str(row[14].value or ""))
        genres.append(str(row[15].value or ""))
        resumes.append(str(row[16].value or ""))
        casting.append(str(row[17].value or ""))

        titres.append(titre)
        liens.append(lien)

    return titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting


def short_text(text, max_len=140):
    return text if len(text) <= max_len else text[:max_len] + "..."


# ----------- TMDB MULTI -----------
def search_tmdb_multi(query):
    try:
        url = "https://api.themoviedb.org/3/search/movie"
        params = {"api_key": API_KEY, "query": query, "language": "fr-FR"}
        data = requests.get(url, params=params).json()

        results = []
        for film in data.get("results", [])[:5]:
            poster = film.get("poster_path")
            img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""
            results.append({
                "id": film["id"],
                "title": film["title"],
                "year": film.get("release_date", "")[:4],
                "img": img
            })
        return results
    except:
        return []


# ----------- HOME -----------
@app.route("/", methods=["GET"])
def home():

    show_suggest = request.args.get("suggest")

    titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting = load_data()

    query = request.args.get("q", "")
    mode = request.args.get("mode")

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}

    input {
        padding:18px;
        width:90%;
        font-size:20px;
        border-radius:14px;
        border:none;
    }

    .card {
        background:#1c1c1c;
        padding:18px;
        border-radius:18px;
        margin:18px auto;
        width:95%;
        max-width:420px;
    }

    img {width:100%; border-radius:14px;}

    .titre a {
        color:#00c3ff;
        font-size:20px;
        font-weight:bold;
        text-decoration:none;
    }

    .btn-row {
        display:flex;
        gap:10px;
        margin-top:14px;
    }

    .btn {
        flex:1;
        padding:14px;
        border-radius:12px;
        text-decoration:none;
        font-size:16px;
        font-weight:bold;
    }

    .allocine {background:#00c3ff; color:black;}
    .new {background:#e50914; color:white;}
    .update {background:#ff9800; color:black;}
    </style>
    """

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film..."><br><br>
            <button name="mode" value="fast">🔎 Recherche</button>
            <button name="mode" value="poster">🎬 Jaquettes</button>
        </form>
        """

    query_clean = query.lower().strip()
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        if fuzz.partial_ratio(query_clean, titre_clean) > 90:
            score += 100

        if score >= 100:
            results.append((score, i))

    query_encoded = urllib.parse.quote(query)

    # ----------- RESULTATS TROUVES -----------
    if results:
        bloc = f"<h2>{len(results)} résultats</h2>"

        for score, i in results:

            poster_html = f'<img src="{jaquettes[i]}">' if jaquettes[i] else ""

            bloc += f"""
            <div class="card">
                {poster_html}
                <div class="titre">
                    <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                        {titres[i]}
                    </a>
                </div>

                <div>{genres[i]}</div>
                <div>{casting[i]}</div>

                <div class="btn-row">
                    <a class="btn allocine" href="{liens[i]}" target="_blank">Allociné</a>
                    <a class="btn update" href="/suggest_update/{i}?q={query_encoded}">🔄 Corriger</a>
                </div>
            </div>
            """

        bloc += f"""
        <div class="card">
            <div class="btn-row">
                <a class="btn allocine" href="/?q={query_encoded}&mode={mode}&suggest=1">
                    🔎 Suggestions TMDB
                </a>
            </div>
        </div>
        """

        # ----------- SUGGESTIONS EN PLUS -----------
        if show_suggest:
            results_tmdb = search_tmdb_multi(query)

            for score, i in results:
                results_tmdb += search_tmdb_multi(titres[i])

            unique = {film["id"]: film for film in results_tmdb}

            bloc += "<h2>Suggestions TMDB</h2>"

            for film in unique.values():
                bloc += f"""
                <div class="card">
                    <img src="{film['img']}">
                    <div>{film['title']} ({film['year']})</div>

                    <div class="btn-row">
                        <a class="btn new" href="/add/{film['id']}">➕ Ajouter</a>
                        <a class="btn update" href="/confirm_update/{i}/{film['id']}">🔄 Update</a>
                    </div>
                </div>
                """

        return html + bloc

    # ----------- AUCUN RESULTAT -----------
    results_tmdb = search_tmdb_multi(query)

    bloc = "<h2>❌ Inexistant</h2>"

    for film in results_tmdb:
        bloc += f"""
        <div class="card">
            <img src="{film['img']}">
            <div>{film['title']} ({film['year']})</div>

            <div class="btn-row">
                <a class="btn new" href="/add/{film['id']}">➕ Ajouter</a>
            </div>
        </div>
        """

    return html + bloc


# ----------- CHOIX UPDATE -----------
@app.route("/suggest_update/<int:index>")
def suggest_update(index):

    query = request.args.get("q", "")
    results_tmdb = search_tmdb_multi(query)

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {background:#111;color:white;text-align:center;font-family:Arial;padding:15px;}
    .card {background:#1c1c1c;padding:15px;border-radius:16px;margin:15px auto;width:95%;max-width:420px;}
    img {width:100%;border-radius:12px;}
    .titre {color:#00c3ff;font-size:18px;font-weight:bold;margin-top:10px;}
    .btn {display:block;margin-top:10px;padding:12px;border-radius:10px;text-decoration:none;font-weight:bold;}
    .update {background:#ff9800;color:black;}
    </style>

    <h2>🎯 Choisir le bon film</h2>
    """

    for film in results_tmdb:
        html += f"""
        <div class="card">
            <img src="{film['img']}">
            <div class="titre">{film['title']} ({film['year']})</div>
            <a class="btn update" href="/confirm_update/{index}/{film['id']}">
                🔄 Mettre à jour TMDB
            </a>
        </div>
        """

    return html


# ----------- UPDATE -----------
@app.route("/confirm_update/<int:index>/<int:tmdb_id>")
def confirm_update(index, tmdb_id):

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {
        "api_key": API_KEY,
        "language": "fr-FR",
        "append_to_response": "credits"
    }

    data = requests.get(url, params=params).json()

    wb = load_workbook("films.xlsx")
    ws = wb.active

    row_excel = index + 2

    ws.cell(row=row_excel, column=COL["TMDB_ID"], value=tmdb_id)
    ws.cell(row=row_excel, column=COL["JAQUETTE"], value=f"https://image.tmdb.org/t/p/w300{data.get('poster_path','')}")
    ws.cell(row=row_excel, column=COL["ANNEE"], value=data.get("release_date", "")[:4])
    ws.cell(row=row_excel, column=COL["GENRES"], value=", ".join([g["name"] for g in data.get("genres", [])]))
    ws.cell(row=row_excel, column=COL["RESUME"], value=data.get("overview", ""))
    ws.cell(row=row_excel, column=COL["CASTING"], value=", ".join([c["name"] for c in data.get("credits", {}).get("cast", [])[:5]]))

    wb.save("films.xlsx")

    return "<h2>✅ Mise à jour OK</h2><a href='/'>Retour</a>"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
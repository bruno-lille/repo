from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz

app = Flask(__name__)

# Charger Excel avec les vrais liens
try:
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres = []
    liens = []
    col_A = []
    col_C = []

    for row in ws.iter_rows(min_row=2):
        cell_A = row[0]   # colonne A
        cell_C = row[2]   # colonne C
        titre_cell = row[5]  # colonne F
        lien_cell = row[6]   # colonne G

        titre = str(titre_cell.value) if titre_cell.value else ""
        a_val = str(cell_A.value) if cell_A.value else ""
        c_val = str(cell_C.value) if cell_C.value else ""

        if lien_cell.hyperlink:
            lien = lien_cell.hyperlink.target
        else:
            lien = str(lien_cell.value) if lien_cell.value else ""

        titres.append(titre)
        liens.append(lien)
        col_A.append(a_val)
        col_C.append(c_val)

except Exception as e:
    titres = []
    liens = []
    col_A = []
    col_C = []
    erreur = str(e)


@app.route("/", methods=["GET"])
def home():
    query = request.args.get("q")
    page = int(request.args.get("page", 1))

    html = """
    <style>
    body {font-family: Arial; background: #111; color: white; text-align: center; padding: 30px;}
    input {padding: 10px; width: 250px; border-radius: 10px; border: none;}
    button {padding: 10px 20px; border-radius: 10px; border: none; background: #e50914; color: white;}
    .card {background: #222; padding: 15px; border-radius: 12px; margin: 10px auto; width: 340px;}
    a {color: #00c3ff; text-decoration: none;}
    .nav {margin-top: 20px;}
    </style>
    """

    if not titres:
        return html + f"<h2>❌ Erreur Excel :</h2><p>{erreur}</p>"

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film...">
            <button type="submit">Rechercher</button>
        </form>
        """

    query_clean = query.lower().strip()

    # MATCH EXACT
    for i, titre in enumerate(titres):
        if query_clean == titre.lower().strip():
            return html + f"""
            <h1>🎬 Résultat exact</h1>
            <div class="card">
                <b>🎬 {titre}</b><br><br>
                📀 {col_C[i]}<br>
                📁 {col_A[i]}<br><br>
                <a href="{liens[i]}" target="_blank">👉 Voir fiche</a>
            </div>
            <br><a href="/">⬅ Retour</a>
            """

    # TRI INTELLIGENT
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()

        score = 0

        if query_clean in titre_clean:
            score += 100

        fuzzy_score = fuzz.partial_ratio(query_clean, titre_clean)

        if fuzzy_score > 85:
            score += fuzzy_score

        if score >= 100:
            results.append((score, i))

    results.sort(reverse=True, key=lambda x: x[0])

    # PAGINATION
    if results:
        par_page = 5
        total = len(results)
        start = (page - 1) * par_page
        end = start + par_page
        page_items = results[start:end]

        bloc = f"<h2>{total} résultats 👇</h2>"

        for score, i in page_items:
            bloc += f"""
            <div class="card">
                <b>🎬 {titres[i]}</b><br><br>
                📀 {col_C[i]}<br>
                📁 {col_A[i]}<br><br>
                <a href="{liens[i]}" target="_blank">👉 Voir fiche</a>
            </div>
            """

        nav = '<div class="nav">'

        if start > 0:
            nav += f'<a href="/?q={query}&page={page-1}">⬅ Précédent</a> '

        if end < total:
            nav += f'<a href="/?q={query}&page={page+1}">Suivant ➡</a>'

        nav += '</div>'

        return html + bloc + nav + '<br><a href="/">⬅ Nouvelle recherche</a>'

    return html + "<h2>Aucun résultat 😢</h2><br><a href='/'>⬅ Retour</a>"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
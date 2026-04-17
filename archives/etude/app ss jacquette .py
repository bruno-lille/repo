from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz

app = Flask(__name__)

# Charger Excel
try:
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres = []
    liens = []
    col_A = []
    col_C = []

    for row in ws.iter_rows(min_row=2):
        cell_A = row[0]
        cell_C = row[2]
        titre_cell = row[5]
        lien_cell = row[6]

        titre = str(titre_cell.value) if titre_cell.value else ""
        a_val = str(cell_A.value) if cell_A.value else ""
        c_val = str(cell_C.value) if cell_C.value else ""

        if lien_cell.hyperlink:
            lien = lien_cell.hyperlink.target
        else:
            lien = str(lien_cell.value) if lien_cell.value else ""

        titres.append(titre)
        liens.append(lien)
        col_A.append(a_val)
        col_C.append(c_val)

except Exception as e:
    titres = []
    liens = []
    col_A = []
    col_C = []
    erreur = str(e)


@app.route("/", methods=["GET"])
def home():
    query = request.args.get("q")

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {
        font-family: Arial;
        background: #111;
        color: white;
        text-align: center;
        padding: 20px;
    }

    h1 { font-size: 28px; }

    input {
        padding: 15px;
        width: 80%;
        font-size: 18px;
        border-radius: 12px;
        border: none;
        margin-bottom: 10px;
    }

    button {
        padding: 15px 25px;
        font-size: 18px;
        border-radius: 12px;
        border: none;
        background: #e50914;
        color: white;
        margin-top: 10px;
    }

    .card {
        background: #222;
        padding: 20px;
        border-radius: 16px;
        margin: 15px auto;
        width: 95%;
        max-width: 400px;
        font-size: 18px;
    }

    .titre {
        font-size: 22px;
        font-weight: bold;
        margin-bottom: 10px;
    }

    .info {
        font-size: 16px;
        color: #ccc;
        margin-bottom: 10px;
    }

    a {
        display: inline-block;
        padding: 12px 20px;
        background: #00c3ff;
        color: black;
        border-radius: 10px;
        font-size: 16px;
        text-decoration: none;
        margin-top: 10px;
    }
    </style>
    """

    if not titres:
        return html + f"<h2>❌ Erreur Excel :</h2><p>{erreur}</p>"

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film...">
            <br>
            <button type="submit">Rechercher</button>
        </form>
        """

    query_clean = query.lower().strip()

    # MATCH EXACT
    for i, titre in enumerate(titres):
        if query_clean == titre.lower().strip():
            return html + f"""
            <h1>🎬 Résultat</h1>
            <div class="card">
                <div class="titre">{titre}</div>
                <div class="info">📀 {col_C[i]}</div>
                <div class="info">📁 {col_A[i]}</div>
                <a href="{liens[i]}" target="_blank">Voir sur AlloCiné</a>
            </div>
            <br><a href="/">Nouvelle recherche</a>
            """

    # TRI INTELLIGENT
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        fuzzy_score = fuzz.partial_ratio(query_clean, titre_clean)

        if fuzzy_score > 85:
            score += fuzzy_score

        if score >= 100:
            results.append((score, i))

    results.sort(reverse=True, key=lambda x: x[0])

    # AFFICHAGE (tout en scroll)
    if results:
        bloc = f"<h2>{len(results)} résultats</h2>"

        for score, i in results:
            bloc += f"""
            <div class="card">
                <div class="titre">{titres[i]}</div>
                <div class="info">📀 {col_C[i]}</div>
                <div class="info">📁 {col_A[i]}</div>
                <a href="{liens[i]}" target="_blank">Voir sur AlloCiné</a>
            </div>
            """

        return html + bloc + '<br><a href="/">Nouvelle recherche</a>'

    return html + "<h2>Aucun résultat</h2><br><a href='/'>Retour</a>"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import urllib.parse
import requests

app = Flask(__name__)

API_KEY = "0458b97834e41cd8c1e0f3364d6703d6"  # ← à remplir plus tard

# ----------- FONCTION JAQUETTE -----------
def get_poster(title):
    if API_KEY == "TA_CLE_TMDB_ICI":
        return ""  # pas de clé → pas d'image

    try:
        url = "https://api.themoviedb.org/3/search/movie"
        params = {"api_key": API_KEY, "query": title, "language": "fr-FR"}
        r = requests.get(url, params=params)
        data = r.json()

        if data["results"]:
            poster = data["results"][0]["poster_path"]
            if poster:
                return f"https://image.tmdb.org/t/p/w300{poster}"
    except:
        pass

    return ""

# ----------- CHARGEMENT EXCEL -----------
try:
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens, col_A, col_C = [], [], [], []

    for row in ws.iter_rows(min_row=2):
        titre = str(row[5].value or "")
        a_val = str(row[0].value or "")
        c_val = str(row[2].value or "")

        if row[6].hyperlink:
            lien = row[6].hyperlink.target
        else:
            lien = str(row[6].value or "")

        titres.append(titre)
        liens.append(lien)
        col_A.append(a_val)
        col_C.append(c_val)

except Exception as e:
    titres, liens, col_A, col_C = [], [], [], []
    erreur = str(e)


@app.route("/", methods=["GET"])
def home():
    query = request.args.get("q")
    mode = request.args.get("mode")  # fast / poster

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:20px;}
    input {padding:15px; width:80%; font-size:18px; border-radius:12px; border:none;}
    button {padding:12px 20px; font-size:16px; border-radius:10px; border:none; margin:5px;}
    .fast {background:#555; color:white;}
    .poster {background:#e50914; color:white;}
    .card {background:#222; padding:15px; border-radius:16px; margin:15px auto; width:95%; max-width:400px;}
    img {width:100%; border-radius:12px; margin-bottom:10px;}
    .titre {font-size:20px; font-weight:bold;}
    .info {color:#ccc;}
    a {display:inline-block; padding:10px; background:#00c3ff; color:black; border-radius:10px; margin-top:10px;}
    </style>
    """

    # PAGE ACCUEIL
    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film..."><br>
            <button class="fast" name="mode" value="fast">🔎 Rapide</button>
            <button class="poster" name="mode" value="poster">🎬 Avec jaquettes</button>
        </form>
        """

    query_clean = query.lower().strip()

    # RECHERCHE
    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        fuzzy_score = fuzz.partial_ratio(query_clean, titre_clean)
        if fuzzy_score > 85:
            score += fuzzy_score

        if score >= 100:
            results.append((score, i))

    results.sort(reverse=True, key=lambda x: x[0])

    # SI RESULTATS
    if results:
        bloc = f"<h2>{len(results)} résultats</h2>"

        for score, i in results:

            poster_html = ""
            if mode == "poster":
                poster_url = get_poster(titres[i])
                if poster_url:
                    poster_html = f'<img src="{poster_url}">'

            bloc += f"""
            <div class="card">
                {poster_html}
                <div class="titre">{titres[i]}</div>
                <div class="info">📀 {col_C[i]}</div>
                <div class="info">📁 {col_A[i]}</div>
                <a href="{liens[i]}" target="_blank">Voir sur Allociné</a>
            </div>
            """

        return html + bloc + '<br><a href="/">Nouvelle recherche</a>'

    # SI AUCUN RESULTAT → GOOGLE
    query_encoded = urllib.parse.quote(query)
    google_link = f"https://www.google.com/search?q={query_encoded}+allocine+fichefilm"

    return html + f"""
    <h2>❌ Aucun résultat dans ta liste</h2>

    <a href="{google_link}" target="_blank">
    🎬 Voir fiche film Allociné
    </a>

    <br><br>
    <a href="/">Retour</a>
    """


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
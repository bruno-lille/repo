from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import urllib.parse
import requests
import re

app = Flask(__name__)

API_KEY = "0458b97834e41cd8c1e0f3364d6703d6"

COL = {
    "EMPLACEMENT": 1,
    "TYPE": 3,
    "TITRE": 6,
    "ALLOCINE": 7,
    "TMDB_ID": 13,
    "JAQUETTE": 14,
    "ANNEE": 15,
    "GENRES": 16,
    "RESUME": 17,
    "CASTING": 18
}

# ----------- STYLE -----------
def style():
    return """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:15px;}
    input {padding:14px; width:85%; font-size:16px; border-radius:10px; border:none; text-align:center;}

    .card {background:#1c1c1c; padding:15px; border-radius:16px; margin:15px auto; width:95%; max-width:420px;}
    img {width:100%; border-radius:12px;}

    .titre a {color:#00c3ff; font-size:18px; font-weight:bold; text-decoration:none;}

    .btn-row {display:flex; gap:8px; margin-top:12px;}
    .btn {flex:1; padding:12px; border-radius:10px; text-decoration:none; font-weight:bold;}

    .allocine {background:#00c3ff; color:black;}
    .new {background:#e50914; color:white;}
    .update {background:#ff9800; color:black;}
    .retour {background:#444; color:white;}
    </style>
    """

# ----------- LOAD DATA -----------
def load_data():
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens = [], []
    col_A, col_C = [], []
    tmdb_ids, jaquettes = [], []
    annees, genres = [], []
    resumes, casting = [], []

    for row in ws.iter_rows(min_row=2):

        titre = str(row[COL["TITRE"] - 1].value or "").strip()
        lien = row[6].hyperlink.target if row[6].hyperlink else str(row[6].value or "")

        col_A.append(str(row[0].value or ""))
        col_C.append(str(row[2].value or ""))

        tmdb_ids.append(str(row[12].value or ""))
        jaquettes.append(str(row[13].value or ""))
        annees.append(str(row[14].value or ""))
        genres.append(str(row[15].value or ""))
        resumes.append(str(row[16].value or ""))
        casting.append(str(row[17].value or ""))

        titres.append(titre)
        liens.append(lien)

    return titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting


# ----------- TMDB SEARCH -----------
def search_tmdb(query):
    try:
        url = "https://api.themoviedb.org/3/search/movie"
        params = {"api_key": API_KEY, "query": query, "language": "fr-FR"}
        data = requests.get(url, params=params).json()

        results = []
        for film in data.get("results", [])[:5]:
            poster = film.get("poster_path")
            img = f"https://image.tmdb.org/t/p/w300{poster}" if poster else ""
            results.append({
                "id": film["id"],
                "title": film["title"],
                "year": film.get("release_date", "")[:4],
                "img": img
            })
        return results
    except:
        return []


# ----------- HOME -----------
@app.route("/")
def home():

    titres, liens, col_A, col_C, tmdb_ids, jaquettes, annees, genres, resumes, casting = load_data()

    query = request.args.get("q", "")
    show_suggest = request.args.get("suggest")

    html = style()

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film..."><br><br>
            <button class="btn allocine">🔎 Rechercher</button>
        </form>
        """

    query_clean = query.lower()
    results = []

    for i, titre in enumerate(titres):
        if query_clean in titre.lower() or fuzz.partial_ratio(query_clean, titre.lower()) > 90:
            results.append(i)

    query_encoded = urllib.parse.quote(query)

    # ----------- RESULTATS -----------
    if results:

        bloc = f"<h2>{len(results)} résultats</h2>"

        # 🔝 boutons haut
        bloc += f"""
        <div class="card">
            <div class="btn-row">
                <a class="btn retour" href="/">⬅️ Retour</a>
                <a class="btn new" href="/">❌ Annuler</a>
            </div>
        </div>
        """

        for i in results:

            poster = ""
            if jaquettes[i]:
                poster = f"""
                <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                    <img src="{jaquettes[i]}">
                </a>
                """

            bloc += f"""
            <div class="card">
                {poster}

                <div class="titre">
                    <a href="https://www.themoviedb.org/movie/{tmdb_ids[i]}" target="_blank">
                        {titres[i]}
                    </a>
                </div>

                <div>{genres[i]}</div>
                <div>{casting[i]}</div>

                <div class="btn-row">
                    <a class="btn allocine" href="{liens[i]}" target="_blank">Allociné</a>
                    <a class="btn update" href="/suggest_update/{i}?q={query_encoded}">🔄 Corriger</a>
                </div>
            </div>
            """

        bloc += f"""
        <div class="card">
            <a class="btn allocine" href="/?q={query_encoded}&suggest=1">
                🔎 Suggestions TMDB
            </a>
        </div>
        """

        return html + bloc

    # ----------- SUGGESTIONS (SANS CORRIGER) -----------
    suggestions = search_tmdb(query)

    bloc = "<h2>❌ Inexistant</h2>"

    for film in suggestions:
        bloc += f"""
        <div class="card">
            <img src="{film['img']}">
            <div>{film['title']} ({film['year']})</div>

            <div class="btn-row">
                <a class="btn new" href="/add/{film['id']}">➕ Ajouter</a>
            </div>
        </div>
        """

    return html + bloc


# ----------- PAGE AJOUT -----------
@app.route("/add/<int:tmdb_id>")
def add_movie(tmdb_id):

    url = f"https://api.themoviedb.org/3/movie/{tmdb_id}"
    params = {"api_key": API_KEY, "language": "fr-FR", "append_to_response": "credits"}
    d = requests.get(url, params=params).json()

    title = d.get("title", "")
    year = d.get("release_date", "")[:4]
    genres = ", ".join([g["name"] for g in d.get("genres", [])])
    overview = d.get("overview", "")
    poster = f"https://image.tmdb.org/t/p/w300{d.get('poster_path','')}"
    cast = ", ".join([a["name"] for a in d.get("credits", {}).get("cast", [])[:5]])

    return f"""
    {style()}

    <h2>➕ Ajouter</h2>

    <img src="{poster}">

    <h3>{title} ({year})</h3>
    <p>{genres}</p>
    <p>{cast}</p>

    <form method="post" action="/confirm_add">
        <input type="hidden" name="tmdb_id" value="{tmdb_id}">
        <input type="hidden" name="title" value="{title}">
        <input type="hidden" name="year" value="{year}">
        <input type="hidden" name="genres" value="{genres}">
        <input type="hidden" name="overview" value="{overview}">
        <input type="hidden" name="cast" value="{cast}">
        <input type="hidden" name="poster" value="{poster}">

        <input name="emplacement" placeholder="Emplacement"><br><br>
        <input name="allocine" placeholder="Lien Allociné"><br><br>

        <button class="btn new">✅ Ajouter</button>
    </form>
    """


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
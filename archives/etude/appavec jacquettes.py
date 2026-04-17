from flask import Flask, request
from openpyxl import load_workbook
from rapidfuzz import fuzz
import requests
import urllib.parse

app = Flask(__name__)

API_KEY = "TA_CLE_TMDB_ICI"  # ← remplace ici

def get_poster(title):
    try:
        url = f"https://api.themoviedb.org/3/search/movie"
        params = {
            "api_key": API_KEY,
            "query": title,
            "language": "fr-FR"
        }
        r = requests.get(url, params=params)
        data = r.json()

        if data["results"]:
            poster_path = data["results"][0]["poster_path"]
            if poster_path:
                return f"https://image.tmdb.org/t/p/w300{poster_path}"
    except:
        pass

    return "https://via.placeholder.com/300x450?text=No+Image"


# Charger Excel
try:
    wb = load_workbook("films.xlsx")
    ws = wb.active

    titres, liens, col_A, col_C = [], [], [], []

    for row in ws.iter_rows(min_row=2):
        titre = str(row[5].value or "")
        a_val = str(row[0].value or "")
        c_val = str(row[2].value or "")

        if row[6].hyperlink:
            lien = row[6].hyperlink.target
        else:
            lien = str(row[6].value or "")

        titres.append(titre)
        liens.append(lien)
        col_A.append(a_val)
        col_C.append(c_val)

except Exception as e:
    titres, liens, col_A, col_C = [], [], [], []
    erreur = str(e)


@app.route("/", methods=["GET"])
def home():
    query = request.args.get("q")

    html = """
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
    body {font-family: Arial; background:#111; color:white; text-align:center; padding:20px;}
    input {padding:15px; width:80%; font-size:18px; border-radius:12px; border:none;}
    button {padding:15px; font-size:18px; border-radius:12px; border:none; background:#e50914; color:white;}
    .card {background:#222; padding:15px; border-radius:16px; margin:15px auto; width:95%; max-width:400px;}
    img {width:100%; border-radius:12px;}
    .titre {font-size:20px; font-weight:bold;}
    .info {color:#ccc;}
    a {display:inline-block; padding:10px; background:#00c3ff; color:black; border-radius:10px; margin-top:10px;}
    </style>
    """

    if not query:
        return html + """
        <h1>🎬 Ma vidéothèque</h1>
        <form>
            <input name="q" placeholder="Tape un film...">
            <br><button type="submit">Rechercher</button>
        </form>
        """

    query_clean = query.lower().strip()

    results = []

    for i, titre in enumerate(titres):
        titre_clean = titre.lower().strip()
        score = 0

        if query_clean in titre_clean:
            score += 100

        fuzzy_score = fuzz.partial_ratio(query_clean, titre_clean)
        if fuzzy_score > 85:
            score += fuzzy_score

        if score >= 100:
            results.append((score, i))

    results.sort(reverse=True, key=lambda x: x[0])

    bloc = f"<h2>{len(results)} résultats</h2>"

    for score, i in results:
        poster = get_poster(titres[i])

        bloc += f"""
        <div class="card">
            <img src="{poster}">
            <div class="titre">{titres[i]}</div>
            <div class="info">📀 {col_C[i]}</div>
            <div class="info">📁 {col_A[i]}</div>
            <a href="{liens[i]}" target="_blank">Voir sur AlloCiné</a>
        </div>
        """

    return html + bloc + '<br><a href="/">Retour</a>'


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)